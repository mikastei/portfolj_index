from pathlib import Path
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook

from src.dashboard_workbook import build_dashboard_workbook


def _write_source_workbook(path: Path) -> None:
    with pd.ExcelWriter(path) as writer:
        pd.DataFrame(
            [
                {
                    "Series_ID": "PORT_ALPHA_REAL",
                    "Display_Name": "Alpha Real",
                    "Series_Type": "PORT",
                    "Portfolio_Name": "Alpha",
                    "Variant": "REAL",
                    "Period": "YTD",
                    "Start_Date": "2026-01-01",
                    "End_Date": "2026-03-01",
                    "Obs_Days": 40,
                    "Return_Total": 0.11,
                    "CAGR": 0.14,
                    "Vol": 0.08,
                    "Sharpe": 1.2,
                    "Sortino": 1.5,
                    "Max_DD": -0.05,
                    "Calmar": 2.8,
                },
                {
                    "Series_ID": "PORT_ALPHA_CUR",
                    "Display_Name": "Alpha Current",
                    "Series_Type": "PORT",
                    "Portfolio_Name": "Alpha",
                    "Variant": "CUR",
                    "Period": "YTD",
                    "Start_Date": "2026-01-01",
                    "End_Date": "2026-03-01",
                    "Obs_Days": 40,
                    "Return_Total": 0.10,
                    "CAGR": 0.13,
                    "Vol": 0.09,
                    "Sharpe": 1.1,
                    "Sortino": 1.4,
                    "Max_DD": -0.06,
                    "Calmar": 2.2,
                },
                {
                    "Series_ID": "PORT_BETA_REAL",
                    "Display_Name": "Beta Real",
                    "Series_Type": "PORT",
                    "Portfolio_Name": "Beta",
                    "Variant": "REAL",
                    "Period": "YTD",
                    "Start_Date": "2026-01-01",
                    "End_Date": "2026-03-01",
                    "Obs_Days": 40,
                    "Return_Total": 0.09,
                    "CAGR": 0.12,
                    "Vol": 0.07,
                    "Sharpe": 1.0,
                    "Sortino": 1.3,
                    "Max_DD": -0.04,
                    "Calmar": 3.0,
                },
                {
                    "Series_ID": "BM_GLOBAL",
                    "Display_Name": "Global BM",
                    "Series_Type": "BM",
                    "Portfolio_Name": "",
                    "Variant": "",
                    "Period": "YTD",
                    "Start_Date": "2026-01-01",
                    "End_Date": "2026-03-01",
                    "Obs_Days": 40,
                    "Return_Total": 0.08,
                    "CAGR": 0.1,
                    "Vol": 0.06,
                    "Sharpe": 0.9,
                    "Sortino": 1.1,
                    "Max_DD": -0.03,
                    "Calmar": 3.3,
                },
                {
                    "Series_ID": "BM_SWEDEN",
                    "Display_Name": "Sweden BM",
                    "Series_Type": "BM",
                    "Portfolio_Name": "",
                    "Variant": "",
                    "Period": "YTD",
                    "Start_Date": "2026-01-01",
                    "End_Date": "2026-03-01",
                    "Obs_Days": 40,
                    "Return_Total": 0.06,
                    "CAGR": 0.08,
                    "Vol": 0.05,
                    "Sharpe": 0.85,
                    "Sortino": 1.0,
                    "Max_DD": -0.025,
                    "Calmar": 3.2,
                },
                {
                    "Series_ID": "CAT_ALPHA_DEF",
                    "Display_Name": "Alpha Defensive",
                    "Series_Type": "PORT",
                    "Portfolio_Name": "Alpha",
                    "Variant": "REAL",
                    "Period": "YTD",
                    "Start_Date": "2026-01-01",
                    "End_Date": "2026-03-01",
                    "Obs_Days": 40,
                    "Return_Total": 0.07,
                    "CAGR": 0.09,
                    "Vol": 0.05,
                    "Sharpe": 0.8,
                    "Sortino": 1.0,
                    "Max_DD": -0.02,
                    "Calmar": 4.5,
                },
                {
                    "Series_ID": "CAT_ALPHA_GROWTH",
                    "Display_Name": "Alpha Growth",
                    "Series_Type": "PORT",
                    "Portfolio_Name": "Alpha",
                    "Variant": "REAL",
                    "Period": "YTD",
                    "Start_Date": "2026-01-01",
                    "End_Date": "2026-03-01",
                    "Obs_Days": 40,
                    "Return_Total": 0.13,
                    "CAGR": 0.16,
                    "Vol": 0.10,
                    "Sharpe": 1.25,
                    "Sortino": 1.55,
                    "Max_DD": -0.045,
                    "Calmar": 3.6,
                },
            ]
        ).to_excel(writer, sheet_name="KPI_Summary", index=False)
        pd.DataFrame(
            [
                {
                    "Series_ID": "PORT_ALPHA_REAL",
                    "Display_Name": "Alpha Real",
                    "Series_Type": "PORT",
                    "Portfolio_Name": "Alpha",
                    "Variant": "REAL",
                    "30D": 0.01,
                    "YTD": 0.11,
                    "1Y": 0.2,
                    "Since_Start": 0.35,
                },
                {
                    "Series_ID": "PORT_BETA_REAL",
                    "Display_Name": "Beta Real",
                    "Series_Type": "PORT",
                    "Portfolio_Name": "Beta",
                    "Variant": "REAL",
                    "30D": 0.02,
                    "YTD": 0.09,
                    "1Y": 0.16,
                    "Since_Start": 0.3,
                },
                {
                    "Series_ID": "BM_GLOBAL",
                    "Display_Name": "Global BM",
                    "Series_Type": "BM",
                    "Portfolio_Name": "",
                    "Variant": "",
                    "30D": 0.015,
                    "YTD": 0.08,
                    "1Y": 0.14,
                    "Since_Start": 0.25,
                },
                {
                    "Series_ID": "BM_SWEDEN",
                    "Display_Name": "Sweden BM",
                    "Series_Type": "BM",
                    "Portfolio_Name": "",
                    "Variant": "",
                    "30D": 0.012,
                    "YTD": 0.06,
                    "1Y": 0.11,
                    "Since_Start": 0.20,
                },
                {
                    "Series_ID": "CAT_ALPHA_DEF",
                    "Display_Name": "Alpha Defensive",
                    "Series_Type": "PORT",
                    "Portfolio_Name": "Alpha",
                    "Variant": "REAL",
                    "30D": 0.008,
                    "YTD": 0.07,
                    "1Y": 0.12,
                    "Since_Start": 0.18,
                },
                {
                    "Series_ID": "CAT_ALPHA_GROWTH",
                    "Display_Name": "Alpha Growth",
                    "Series_Type": "PORT",
                    "Portfolio_Name": "Alpha",
                    "Variant": "REAL",
                    "30D": 0.018,
                    "YTD": 0.13,
                    "1Y": 0.21,
                    "Since_Start": 0.29,
                },
            ]
        ).to_excel(
            writer,
            sheet_name="Period_Returns",
            index=False,
        )
        pd.DataFrame(
            [
                {
                    "Date": "2026-02-28",
                    "PORT_ALPHA_REAL": 99.0,
                    "PORT_BETA_REAL": 98.0,
                    "BM_GLOBAL": 97.0,
                    "BM_SWEDEN": 96.5,
                    "CAT_ALPHA_DEF": 99.5,
                    "CAT_ALPHA_GROWTH": 98.8,
                },
                {
                    "Date": "2026-03-01",
                    "PORT_ALPHA_REAL": 100.0,
                    "PORT_BETA_REAL": 99.0,
                    "BM_GLOBAL": 98.0,
                    "BM_SWEDEN": 97.3,
                    "CAT_ALPHA_DEF": 100.0,
                    "CAT_ALPHA_GROWTH": 100.0,
                },
            ]
        ).to_excel(
            writer,
            sheet_name="Chart_IDX_Wide",
            index=False,
        )
        pd.DataFrame(
            [
                {
                    "Date": "2026-02-28",
                    "PORT_ALPHA_REAL": -0.01,
                    "PORT_BETA_REAL": -0.02,
                    "BM_GLOBAL": -0.03,
                    "BM_SWEDEN": -0.035,
                    "CAT_ALPHA_DEF": -0.01,
                    "CAT_ALPHA_GROWTH": -0.015,
                },
                {
                    "Date": "2026-03-01",
                    "PORT_ALPHA_REAL": 0.0,
                    "PORT_BETA_REAL": -0.01,
                    "BM_GLOBAL": -0.02,
                    "BM_SWEDEN": -0.025,
                    "CAT_ALPHA_DEF": 0.0,
                    "CAT_ALPHA_GROWTH": 0.0,
                },
            ]
        ).to_excel(
            writer,
            sheet_name="Chart_DD_Wide",
            index=False,
        )
        pd.DataFrame(
            [
                {
                    "Portfolio_Name": "Alpha",
                    "Series_ID": "PORT_ALPHA_REAL",
                    "Variant": "REAL",
                    "Display_Name": "Alpha Fund A",
                    "Yahoo_Ticker": "AAA",
                    "Weight": 0.60,
                    "Weight_Source": "Andel",
                },
                {
                    "Portfolio_Name": "Alpha",
                    "Series_ID": "PORT_ALPHA_REAL",
                    "Variant": "REAL",
                    "Display_Name": "Alpha Fund B",
                    "Yahoo_Ticker": "BBB",
                    "Weight": 0.40,
                    "Weight_Source": "Andel",
                },
                {
                    "Portfolio_Name": "Alpha",
                    "Series_ID": "PORT_ALPHA_CUR",
                    "Variant": "CUR",
                    "Display_Name": "Alpha Fund A",
                    "Yahoo_Ticker": "AAA",
                    "Weight": 1.00,
                    "Weight_Source": "AndelP",
                },
                {
                    "Portfolio_Name": "Beta",
                    "Series_ID": "PORT_BETA_REAL",
                    "Variant": "REAL",
                    "Display_Name": "Beta Fund C",
                    "Yahoo_Ticker": "CCC",
                    "Weight": 0.55,
                    "Weight_Source": "Andel",
                },
                {
                    "Portfolio_Name": "Beta",
                    "Series_ID": "PORT_BETA_REAL",
                    "Variant": "REAL",
                    "Display_Name": "Beta Fund D",
                    "Yahoo_Ticker": "DDD",
                    "Weight": 0.45,
                    "Weight_Source": "Andel",
                },
            ]
        ).to_excel(writer, sheet_name="Allocation_Snapshot", index=False)
        pd.DataFrame(
            [
                {"Config_Key": "default_portfolio", "Config_Value": "Beta"},
                {"Config_Key": "default_period", "Config_Value": "YTD"},
            ]
        ).to_excel(writer, sheet_name="Dashboard_Config", index=False)
        pd.DataFrame(
            [
                {
                    "Created_At": "2026-03-13 08:00:00",
                    "Source_Output_File": "portfolio_output_timeseries.xlsx",
                    "Date_Min": "2025-01-01",
                    "Date_Max": "2026-03-01",
                    "Number_Of_Analysis_Series": 7,
                    "Number_Of_Portfolios": 2,
                    "Number_Of_Benchmarks": 2,
                    "RF_RATE_ANNUAL": 0.0,
                    "TRADING_DAYS_PER_YEAR": 252,
                }
            ]
        ).to_excel(
            writer,
            sheet_name="Build_Info",
            index=False,
        )


def test_dashboard_workbook_builds_control_and_calc_contracts(tmp_path: Path) -> None:
    source_path = tmp_path / "portfolio_dashboard_data.xlsx"
    output_path = tmp_path / "portfolio_dashboard.xlsx"
    _write_source_workbook(source_path)

    build_dashboard_workbook(source_path, output_path)

    assert output_path.exists()

    wb = load_workbook(output_path, data_only=False)
    control_ws = wb["Control"]
    calc_ws = wb["Calc_Main"]
    calc_category_ws = wb["Calc_Category"]
    overview_ws = wb["Overview"]
    performance_ws = wb["Performance"]
    structure_ws = wb["Structure"]
    category_ws = wb["Category"]

    assert control_ws["A2"].value == "Reference view only. Edit main controls on Overview and category controls on Category."
    assert control_ws["A3"].value == "Open Overview controls"
    assert control_ws["B3"].value == "Open Category controls"
    assert control_ws["B6"].value == "=Control_Selected_Portfolio"
    assert control_ws["B7"].value == "=Control_Selected_Variant"
    assert control_ws["B8"].value == "=Control_Selected_Period"

    defined_names = wb.defined_names
    assert defined_names["Control_Selected_Portfolio"].attr_text == "'Overview'!$A$5:$A$5"
    assert defined_names["Control_Selected_Variant"].attr_text == "'Overview'!$B$5:$B$5"
    assert defined_names["Control_Selected_Period"].attr_text == "'Overview'!$C$5:$C$5"
    assert "Control_Selected_Compare_3" not in defined_names
    assert defined_names["List_Main_Portfolio"].attr_text.startswith("'Lists'!$A$")
    assert defined_names["Lookup_Main_Series_ID"].attr_text.startswith("'Lists'!$D$")
    assert defined_names["Lookup_Main_Series_Key"].attr_text.startswith("'Lists'!$E$")

    assert len(control_ws.data_validations.dataValidation) == 0

    overview_validations = {str(validation.sqref): validation.formula1 for validation in overview_ws.data_validations.dataValidation}
    assert overview_validations["A5"] == "=List_Main_Portfolio"
    assert overview_validations["B5"] == "=List_Variant"
    assert overview_validations["C5"] == "=List_Period"
    assert overview_validations["D5"] == "=List_Compare_Series"
    assert overview_validations["E5"] == "=List_Compare_Series"
    assert "F5" not in overview_validations

    assert category_ws["A5"].value == "Alpha"
    assert category_ws["B5"].value == "REAL"
    assert category_ws["C5"].value == "YTD"
    assert category_ws["D5"].value == "Alpha Defensive"
    assert category_ws["E5"].value == "Global BM"
    assert defined_names["Control_Selected_Category_Portfolio"].attr_text == "'Category'!$A$5:$A$5"
    assert defined_names["Control_Selected_Category_Variant"].attr_text == "'Category'!$B$5:$B$5"
    assert defined_names["Control_Selected_Category_Chart"].attr_text == "'Category'!$D$5:$D$5"
    assert defined_names["Control_Selected_Category_Benchmark"].attr_text == "'Category'!$E$5:$E$5"
    assert defined_names["List_Category_Portfolio"].attr_text.startswith("'Lists'!$A$")
    assert defined_names["Lookup_Category_Series_ID"].attr_text.startswith("'Lists'!$C$")
    assert defined_names["Lookup_Category_Key"].attr_text.startswith("'Lists'!$E$")
    assert defined_names["Lookup_Category_Row_Index"].attr_text.startswith("'Lists'!$F$")
    assert defined_names["Lookup_Category_Portfolio_Row_Key"].attr_text.startswith("'Lists'!$G$")
    assert defined_names["List_Category_Benchmark"].attr_text.startswith("'Lists'!$A$")
    assert defined_names["Lookup_Benchmark_Display_Name"].attr_text.startswith("'Lists'!$A$")
    assert defined_names["Lookup_Benchmark_Series_ID"].attr_text.startswith("'Lists'!$B$")
    category_validations = {str(validation.sqref): validation.formula1 for validation in category_ws.data_validations.dataValidation}
    assert category_validations["A5"] == "=List_Category_Portfolio"
    assert category_validations["B5"] == '="REAL"'
    assert category_validations["C5"] == "=List_Period"
    assert category_validations["D5"] == "=List_Category_Available_Series"
    assert category_validations["E5"] == "=List_Category_Benchmark_Series"
    assert "F5" not in category_validations

    assert calc_ws["B6"].value.startswith("=IFERROR(INDEX(")
    assert "Control_Selected_Portfolio" in calc_ws["B6"].value
    assert "Lookup_Main_Series_Display_Name" in calc_ws["B6"].value
    assert calc_ws["B7"].value.startswith("=IFERROR(INDEX(")
    assert "Lookup_Main_Series_ID" in calc_ws["B7"].value
    assert calc_ws["B10"].value.startswith("=IF(")
    assert "Lookup_Compare_Series_ID" in calc_ws["B10"].value
    assert wb.defined_names["Calc_Main_Overview_Base"].attr_text == "'Calc_Main'!$D$6:$F$8"
    assert wb.defined_names["Calc_Main_Performance_Selected"].attr_text == "'Calc_Main'!$H$18:$J$20"
    assert wb.defined_names["Calc_Main_Performance_IDX_Data"].attr_text.startswith("'Calc_Main'!$L$17:$O$")
    assert wb.defined_names["Calc_Main_Performance_DD_Data"].attr_text.startswith("'Calc_Main'!$Q$17:$T$")
    assert wb.defined_names["Calc_Main_Primary_Display_Name"].attr_text == "'Calc_Main'!$B$6:$B$6"
    assert wb.defined_names["Calc_Main_Primary_Series_ID"].attr_text == "'Calc_Main'!$B$7:$B$7"
    assert wb.defined_names["Calc_Main_Selected_Period"].attr_text == "'Calc_Main'!$B$8:$B$8"
    assert calc_ws["I18"].value == "=B6"
    assert calc_ws["J18"].value == "=B7"
    assert "MATCH($J$18,Source_Chart_IDX_Wide!$1:$1,0)" in calc_ws["M18"].value
    assert "MATCH($J$18,Source_Chart_DD_Wide!$1:$1,0)" in calc_ws["R18"].value
    assert calc_ws["B6"].value == '=IFERROR(INDEX(Lookup_Main_Series_Display_Name,MATCH(Control_Selected_Portfolio&"|"&Control_Selected_Variant,Lookup_Main_Series_Key,0)),"")'
    assert calc_ws["B7"].value == '=IFERROR(INDEX(Lookup_Main_Series_ID,MATCH(Control_Selected_Portfolio&"|"&Control_Selected_Variant,Lookup_Main_Series_Key,0)),"")'
    assert calc_ws["A13"].value is None
    assert calc_ws["B13"].value is None

    assert wb["Source_KPI_Summary"].sheet_state == "hidden"
    assert wb["Source_Period_Returns"].sheet_state == "hidden"
    assert wb["Source_Build_Info"].sheet_state == "hidden"
    assert wb["Source_Chart_IDX_Wide"].sheet_state == "hidden"
    assert wb["Source_Chart_DD_Wide"].sheet_state == "hidden"
    assert wb["Source_Allocation_Snapshot"].sheet_state == "hidden"
    assert overview_ws["A4"].value == "Primary portfolio"
    assert overview_ws["A5"].value == "Beta"
    assert overview_ws["B5"].value == "REAL"
    assert overview_ws["C5"].value == "YTD"
    assert overview_ws["G5"].value == '=IFERROR(INDEX(Source_Build_Date_Max,1),"")'
    assert overview_ws["F4"].value is None
    assert overview_ws["F5"].value is None
    assert overview_ws["A6"].value == "Edit selections here. Performance and Structure mirror the same main controls."
    assert overview_ws["B7"].value == "=Calc_Main_Primary_Display_Name"
    assert "Source_KPI_Return_Total" in overview_ws["B9"].value
    assert "Calc_Main_Primary_Series_ID" in overview_ws["B9"].value
    assert overview_ws["D9"].value == '=IF($F9="","",\'Calc_Main\'!D6)'
    assert overview_ws["F10"].value == "='Calc_Main'!F7"
    assert overview_ws["I9"].value.startswith('=IF($F9="","",IFERROR(INDEX(')
    assert "Source_KPI_Return_Total" in overview_ws["I9"].value
    assert overview_ws["D19"].value.startswith('=IF($C19="","",IFERROR(INDEX(')
    assert "Source_Period_30D" in overview_ws["D19"].value
    assert overview_ws["D12"].value is None
    assert overview_ws["A22"].value is None

    assert performance_ws["A4"].value == "Primary portfolio"
    assert performance_ws["A5"].value == "=Control_Selected_Portfolio"
    assert performance_ws["G4"].value == "Edit controls"
    assert performance_ws["G5"].value == "Open main controls"
    assert performance_ws["F4"].value == "Data through"
    assert performance_ws["F5"].value == '=IFERROR(INDEX(Source_Build_Date_Max,1),"")'
    assert performance_ws["A6"].value == "Main controls are edited on Overview and reused here."
    assert performance_ws["A7"].value == "Active selection"
    assert performance_ws["A9"].value == '=IF($C9="","",\'Calc_Main\'!H18)'
    assert performance_ws["B9"].value == '=IF($C9="","",\'Calc_Main\'!I18)'
    assert performance_ws["C9"].value == "='Calc_Main'!J18"
    assert performance_ws["D9"].value == '=IF($C9="","",Calc_Main_Selected_Period)'
    assert performance_ws["I9"].value.startswith('=IF($H9="","",IFERROR(INDEX(')
    assert "Source_Period_30D" in performance_ws["I9"].value
    assert "Source_KPI_Vol" in performance_ws["M9"].value
    assert performance_ws["B16"].value == '=IF(\'Calc_Main\'!J18="","",\'Calc_Main\'!I18)'
    assert performance_ws["C16"].value == '=IF(\'Calc_Main\'!J19="","",\'Calc_Main\'!I19)'
    assert performance_ws["D16"].value == '=IF(\'Calc_Main\'!J20="","",\'Calc_Main\'!I20)'
    assert performance_ws["E16"].value is None
    assert performance_ws["A17"].value == "='Calc_Main'!L18"
    assert performance_ws["B17"].value == "='Calc_Main'!M18"
    assert performance_ws["G17"].value == "='Calc_Main'!Q18"
    assert performance_ws["A12"].value is None
    assert len(performance_ws._charts) == 2
    with ZipFile(output_path) as archive:
        chart_xml = {
            name: archive.read(name).decode("utf-8")
            for name in archive.namelist()
            if name.startswith("xl/charts/chart")
        }
    performance_drawdown_chart = next(
        xml
        for xml in chart_xml.values()
        if "<a:t>Drawdown</a:t>" in xml and "'Performance'!" in xml
    )
    assert "'Performance'!$G$17:$G$" in performance_drawdown_chart

    assert wb.defined_names["Calc_Main_Structure_Status"].attr_text == "'Calc_Main'!$X$8:$X$8"
    assert wb.defined_names["Calc_Main_Structure_Variant"].attr_text == "'Calc_Main'!$X$7:$X$7"
    assert wb.defined_names["Calc_Main_Structure_Table"].attr_text == "'Calc_Main'!$X$18:$AA$42"
    assert wb.defined_names["Calc_Main_Structure_Top_Data"].attr_text == "'Calc_Main'!$X$45:$Y$55"
    assert calc_ws["X8"].value.startswith('=IF($X$6=""')
    assert calc_ws["X7"].value.startswith("=IF(COUNTIFS(")
    assert "Allocation snapshot available" in calc_ws["X8"].value
    assert "REAL allocation snapshot missing in source data" in calc_ws["X8"].value
    assert calc_ws["X9"].value.startswith("=COUNTIFS(")
    assert "Source_Allocation_Snapshot!$A$2:$A$" in calc_ws["X9"].value
    assert calc_ws["X10"].value.startswith("=SUMIFS(")
    assert calc_ws["W18"].value.startswith('=IFERROR(MATCH(')
    assert "Source_Allocation_Lookup_Key" in calc_ws["W18"].value
    assert calc_ws["X18"].value.startswith('=IF($W18="","",INDEX(Source_Allocation_Snapshot!$D$2:$D$')
    assert calc_ws["Y18"].value.startswith('=IF($W18="","",INDEX(Source_Allocation_Snapshot!$F$2:$F$')

    assert calc_category_ws["B6"].value == "=Control_Selected_Category_Portfolio"
    assert calc_category_ws["B7"].value == '=IF(Control_Selected_Category_Variant="","REAL","REAL")'
    assert calc_category_ws["B8"].value == "=Control_Selected_Category_Period"
    assert calc_category_ws["B9"].value == "=Control_Selected_Category_Chart"
    assert calc_category_ws["B10"].value.startswith('=IF(Control_Selected_Category_Chart="","",IFERROR(INDEX(')
    assert "Lookup_Category_Series_ID" in calc_category_ws["B10"].value
    assert "Lookup_Category_Key" in calc_category_ws["B10"].value
    assert calc_category_ws["B11"].value == "=Control_Selected_Category_Benchmark"
    assert "Lookup_Benchmark_Series_ID" in calc_category_ws["B12"].value
    assert calc_category_ws["K6"].value.startswith('=IFERROR(INDEX(Lookup_Category_Display_Name,MATCH(')
    assert calc_category_ws["L6"].value.startswith('=IFERROR(INDEX(Lookup_Category_Series_ID,MATCH(')
    assert calc_category_ws["N6"].value.startswith('=IFERROR(INDEX(Lookup_Benchmark_Display_Name,')
    assert calc_category_ws["O6"].value.startswith('=IF($N6="","",IFERROR(INDEX(Lookup_Benchmark_Series_ID,')
    assert "Lookup_Category_Portfolio_Row_Key" in calc_category_ws["J6"].value
    assert 'MATCH($B$6&"|"&ROWS($K$6:K6)' in calc_category_ws["J6"].value
    assert wb.defined_names["Calc_Category_Selected_Portfolio"].attr_text == "'Calc_Category'!$B$6:$B$6"
    assert wb.defined_names["Calc_Category_Selected_Period"].attr_text == "'Calc_Category'!$B$8:$B$8"
    assert wb.defined_names["Calc_Category_Available_Count"].attr_text == "'Calc_Category'!$B$13:$B$13"
    assert wb.defined_names["Calc_Category_Status"].attr_text == "'Calc_Category'!$B$14:$B$14"
    assert wb.defined_names["Calc_Category_Chart_Selected"].attr_text == "'Calc_Category'!$D$5:$F$7"
    assert wb.defined_names["Calc_Category_Index_Data"].attr_text.startswith("'Calc_Category'!$Q$5:$S$")
    assert wb.defined_names["Calc_Category_Drawdown_Data"].attr_text.startswith("'Calc_Category'!$U$5:$W$")
    assert wb.defined_names["List_Category_Available_Series"].attr_text.startswith("OFFSET('Calc_Category'!$K$6")
    assert wb.defined_names["List_Category_Benchmark_Series"].attr_text.startswith("OFFSET('Calc_Category'!$N$6")
    assert "MATCH($F$6,Source_Chart_IDX_Wide!$1:$1,0)" in calc_category_ws["R6"].value
    assert "MATCH($F$7,Source_Chart_IDX_Wide!$1:$1,0)" in calc_category_ws["S6"].value
    assert "MATCH($F$6,Source_Chart_DD_Wide!$1:$1,0)" in calc_category_ws["V6"].value
    assert "MATCH($F$7,Source_Chart_DD_Wide!$1:$1,0)" in calc_category_ws["W6"].value

    lists_ws = wb["Lists"]
    main_lookup_header_row = next(
        row_index
        for row_index in range(1, lists_ws.max_row + 1)
        if lists_ws.cell(row=row_index, column=1).value == "Portfolio_Name"
        and lists_ws.cell(row=row_index, column=2).value == "Variant"
        and lists_ws.cell(row=row_index, column=3).value == "Display_Name"
        and lists_ws.cell(row=row_index, column=4).value == "Series_ID"
        and lists_ws.cell(row=row_index, column=5).value == "Lookup_Key"
    )
    main_lookup_rows = []
    row_index = main_lookup_header_row + 1
    while lists_ws.cell(row=row_index, column=1).value:
        main_lookup_rows.append(
            tuple(lists_ws.cell(row=row_index, column=column_index).value for column_index in range(1, 6))
        )
        row_index += 1
    assert ("Beta", "REAL", "Beta Real", "PORT_BETA_REAL", "Beta|REAL") in main_lookup_rows
    assert sum(1 for row in main_lookup_rows if row[:2] == ("Beta", "REAL")) == 1

    category_lookup_header_row = next(
        row_index
        for row_index in range(1, lists_ws.max_row + 1)
        if lists_ws.cell(row=row_index, column=1).value == "Portfolio_Name"
        and lists_ws.cell(row=row_index, column=2).value == "Category_Display_Name"
        and lists_ws.cell(row=row_index, column=3).value == "Series_ID"
        and lists_ws.cell(row=row_index, column=4).value == "Variant"
        and lists_ws.cell(row=row_index, column=5).value == "Lookup_Key"
        and lists_ws.cell(row=row_index, column=6).value == "Row_Index"
        and lists_ws.cell(row=row_index, column=7).value == "Portfolio_Row_Key"
    )
    assert lists_ws.cell(row=category_lookup_header_row + 1, column=6).value == 1
    assert lists_ws.cell(row=category_lookup_header_row + 1, column=7).value == "Alpha|1"
    benchmark_lookup_header_row = next(
        row_index
        for row_index in range(1, lists_ws.max_row + 1)
        if lists_ws.cell(row=row_index, column=1).value == "Display_Name"
        and lists_ws.cell(row=row_index, column=2).value == "Series_ID"
        and lists_ws.cell(row=row_index, column=3).value == "Row_Index"
    )
    benchmark_lookup_rows = []
    row_index = benchmark_lookup_header_row + 1
    while lists_ws.cell(row=row_index, column=1).value:
        benchmark_lookup_rows.append(
            tuple(lists_ws.cell(row=row_index, column=column_index).value for column_index in range(1, 4))
        )
        row_index += 1
    assert ("Global BM", "BM_GLOBAL", 1) in benchmark_lookup_rows
    assert ("Sweden BM", "BM_SWEDEN", 2) in benchmark_lookup_rows

    assert wb.defined_names["Source_KPI_Lookup_Key"].attr_text.startswith("'Source_KPI_Summary'!")

    assert structure_ws["A4"].value == "Primary portfolio"
    assert structure_ws["A5"].value == "=Control_Selected_Portfolio"
    assert structure_ws["C4"].value == "Selection status"
    assert structure_ws["D4"].value == "Snapshot as-of"
    assert structure_ws["E4"].value == "Edit controls"
    assert structure_ws["B5"].value == "=Calc_Main_Structure_Variant"
    assert structure_ws["C5"].value == "=Calc_Main_Structure_Status"
    assert structure_ws["D5"].value == '=IFERROR(INDEX(Source_Build_Date_Max,1),"")'
    assert structure_ws["E5"].value == "Open main controls"
    assert structure_ws["A6"].value == "Main controls are edited on Overview and reused here."
    assert structure_ws["C5"].value != '=IF(Control_Selected_Period="","",Control_Selected_Period&" (read-only)")'
    assert structure_ws["B8"].value == "=Calc_Main_Primary_Display_Name"
    assert structure_ws["B9"].value == "=Calc_Main_Structure_Row_Count"
    assert structure_ws["B10"].value == "=Calc_Main_Structure_Weight_Sum"
    assert structure_ws["D9"].value == "='Calc_Main'!X18"
    assert structure_ws["E9"].value == "='Calc_Main'!Y18"
    assert structure_ws["F9"].value == "='Calc_Main'!Z18"
    assert structure_ws["I9"].value == "='Calc_Main'!X46"
    assert structure_ws["J9"].value == "='Calc_Main'!Y46"
    assert wb["Source_Allocation_Snapshot"]["D2"].value == "Alpha Fund A"
    assert wb["Source_Allocation_Snapshot"]["E2"].value == "AAA"
    assert len(structure_ws._charts) == 1

    assert category_ws["A1"].value == "Category"
    assert category_ws["A2"].value == "Category analysis (REAL only) with its own editable controls."
    assert category_ws["A4"].value == "Portfolio"
    assert category_ws["A5"].value == "Alpha"
    assert category_ws["B5"].value == "REAL"
    assert category_ws["C5"].value == "YTD"
    assert category_ws["D4"].value == "Chart category"
    assert category_ws["E4"].value == "Benchmark"
    assert category_ws["A6"].value == "Edit category selections here. This sheet is the only source for category controls."
    assert category_ws["A7"].value == "Category KPI comparison"
    assert category_ws["A8"].value == "Category"
    assert category_ws["A9"].value == '=IF($B9="","",\'Calc_Category\'!K6)'
    assert category_ws["B9"].value == "='Calc_Category'!L6"
    assert category_ws["C9"].value == '=IF($B9="","",Calc_Category_Selected_Period)'
    assert category_ws["A10"].value == '=IF($B10="","",\'Calc_Category\'!K7)'
    assert category_ws["B10"].value == "='Calc_Category'!L7"
    assert "Source_KPI_Return_Total" in category_ws["D9"].value
    assert "Calc_Category_Selected_Period" in category_ws["E9"].value
    assert category_ws["A17"].value == "Index development"
    assert category_ws["A18"].value == "Date"
    assert category_ws["B18"].value == '=IF(\'Calc_Category\'!F6="","",\'Calc_Category\'!E6)'
    assert category_ws["C18"].value == '=IF(\'Calc_Category\'!F7="","",\'Calc_Category\'!E7)'
    assert category_ws["A19"].value == "='Calc_Category'!Q6"
    assert category_ws["B19"].value == "='Calc_Category'!R6"
    assert category_ws["C19"].value == "='Calc_Category'!S6"
    assert category_ws["G17"].value == "Drawdown"
    assert category_ws["G19"].value == "='Calc_Category'!U6"
    assert len(category_ws._charts) == 2
    category_drawdown_chart = next(
        xml
        for xml in chart_xml.values()
        if "<a:t>Category drawdown</a:t>" in xml and "'Category'!" in xml
    )
    assert "'Category'!$G$19:$G$" in category_drawdown_chart


def test_dashboard_workbook_builds_structure_view_from_real_data(tmp_path: Path) -> None:
    source_path = Path("data/portfolio_dashboard_data.xlsx")
    if not source_path.exists():
        return

    output_path = tmp_path / "portfolio_dashboard_real.xlsx"
    build_dashboard_workbook(source_path, output_path)

    assert output_path.exists()

    wb = load_workbook(output_path, data_only=False)
    structure_ws = wb["Structure"]
    calc_ws = wb["Calc_Main"]

    assert structure_ws["A1"].value == "Structure"
    assert structure_ws["A2"].value == "Current allocation snapshot driven by the shared main controls."
    assert structure_ws["D7"].value == "Allocation table"
    assert structure_ws["I7"].value == "Chart helper"
    assert structure_ws["D9"].value == "='Calc_Main'!X18"
    assert structure_ws["I9"].value == "='Calc_Main'!X46"
    assert len(structure_ws._charts) == 1
    assert wb["Source_Allocation_Snapshot"].sheet_state == "hidden"
    assert calc_ws["X9"].value.startswith("=COUNTIFS(")
