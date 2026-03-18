"""Excel input loading."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook

COL_AFFARSDAG = "Aff\u00e4rsdag"
COL_PORTFOLJ = "Portf\u00f6lj"
CRITICAL_FORMULA_COLUMNS = {
    "Transactions": ("Belopp", "Inköpsvärde", "Inkopsvarde"),
}

logger = logging.getLogger(__name__)


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(c).strip() for c in out.columns]
    return out


def _find_table(ws, table_name: str):
    """
    Return openpyxl Table object for a given name across openpyxl versions.
    """
    target = table_name.casefold()

    # Preferred: ws._tables often stores actual Table objects.
    raw_tables = getattr(ws, "_tables", None)
    if raw_tables is not None:
        candidates = raw_tables.values() if isinstance(raw_tables, dict) else raw_tables
        for t in candidates:
            name = getattr(t, "name", None)
            if isinstance(name, str) and name.casefold() == target:
                return t

    # Fallback: ws.tables may be dict-like with values as Table objects or strings.
    tables = getattr(ws, "tables", None)
    if isinstance(tables, dict):
        for v in tables.values():
            name = getattr(v, "name", None)
            if isinstance(name, str) and name.casefold() == target:
                return v
        v = tables.get(table_name)
        if getattr(v, "ref", None) is not None:
            return v
        for key in tables.keys():
            if isinstance(key, str) and key.casefold() == target:
                v = tables[key]
                if getattr(v, "ref", None) is not None:
                    return v

    raise KeyError(f"Could not find Excel table '{table_name}' on sheet '{ws.title}'")


def _raise_on_missing_cached_formulas(
    path: str | Path,
    sheet_name: str,
    table_ref: str,
    table_name: str,
    value_rows,
) -> None:
    critical_columns = CRITICAL_FORMULA_COLUMNS.get(table_name)
    if not critical_columns or not value_rows:
        return

    wb_formula = load_workbook(path, data_only=False, read_only=False)
    try:
        ws_formula = wb_formula[sheet_name]
        formula_rows = ws_formula[table_ref]
        if not formula_rows:
            return

        header = [str(cell.value).strip() if cell.value is not None else "" for cell in formula_rows[0]]
        column_indexes = {
            column_name: idx for idx, column_name in enumerate(header) if column_name in critical_columns
        }
        if not column_indexes:
            return

        issues: list[str] = []
        for table_row_number, (formula_row, value_row) in enumerate(zip(formula_rows[1:], value_rows[1:]), start=1):
            for column_name, idx in column_indexes.items():
                formula_cell = formula_row[idx]
                value_cell = value_row[idx]
                if formula_cell.data_type != "f" or value_cell.value is not None:
                    continue
                issues.append(
                    f"{column_name} {formula_cell.coordinate} (tabellrad {table_row_number}, formel {formula_cell.value})"
                )

        if not issues:
            return

        preview = "; ".join(issues[:5])
        if len(issues) > 5:
            preview = f"{preview}; +{len(issues) - 5} till"
        message = (
            f"Excel table '{table_name}' on sheet '{sheet_name}' contains formula cells without cached values "
            f"in critical columns: {preview}. The workbook is loaded with data_only=True, so openpyxl reads "
            f"these cells as blank/None until the workbook has been recalculated and saved in Excel."
        )
        logger.error(message)
        raise ValueError(message)
    finally:
        wb_formula.close()


def _table_to_dataframe(path: str | Path, table_name: str) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True, read_only=False)
    try:
        for ws in wb.worksheets:
            try:
                table = _find_table(ws, table_name)
            except KeyError:
                continue

            rows = ws[table.ref]
            _raise_on_missing_cached_formulas(path, ws.title, table.ref, table_name, rows)
            raw = [[cell.value for cell in row] for row in rows]
            if not raw:
                return pd.DataFrame()
            header = [str(h).strip() if h is not None else "" for h in raw[0]]
            data = raw[1:]
            return _normalize_columns(pd.DataFrame(data, columns=header))
    finally:
        wb.close()

    raise ValueError(f"Structured table '{table_name}' not found in {path}")


def _validate_columns(df: pd.DataFrame, required: Iterable[str], table_name: str) -> None:
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise ValueError(f"Table '{table_name}' is missing required columns: {missing}")


def load_inputs(path_transaktioner: str, path_fonder: str) -> dict[str, pd.DataFrame]:
    """Load all required structured tables."""
    transactions = _table_to_dataframe(path_transaktioner, "Transactions")
    mapping = _table_to_dataframe(path_transaktioner, "Mapping")
    portfolio_metadata = _table_to_dataframe(path_transaktioner, "Portfolio_Metadata")
    benchmarks = _table_to_dataframe(path_transaktioner, "Benchmarks")
    fondertabell = _table_to_dataframe(path_fonder, "Fondertabell")

    _validate_columns(
        transactions,
        [COL_AFFARSDAG, "ISIN", "Antal", "Kurs", "Valuta", "Transaktionstyp"],
        "Transactions",
    )
    _validate_columns(
        mapping,
        ["ISIN", "Yahoo_Ticker", "Instrument_Type", "Price_Currency", "Category"],
        "Mapping",
    )
    _validate_columns(
        portfolio_metadata,
        ["Portfolio_Name", "Index_Start_Date", "Initial_Index_Value"],
        "Portfolio_Metadata",
    )
    _validate_columns(benchmarks, ["Benchmark_ID", "Yahoo_Ticker", "Include_From_Date"], "Benchmarks")
    _validate_columns(fondertabell, [COL_PORTFOLJ, "Yahoo", "Andel", "AndelP"], "Fondertabell")

    transactions[COL_AFFARSDAG] = pd.to_datetime(transactions[COL_AFFARSDAG], errors="coerce")
    portfolio_metadata["Index_Start_Date"] = pd.to_datetime(
        portfolio_metadata["Index_Start_Date"], errors="coerce"
    )
    benchmarks["Include_From_Date"] = pd.to_datetime(benchmarks["Include_From_Date"], errors="coerce")
    transactions["Antal"] = pd.to_numeric(transactions["Antal"], errors="coerce")
    transactions["Kurs"] = pd.to_numeric(transactions["Kurs"], errors="coerce")
    fondertabell["Andel"] = pd.to_numeric(fondertabell["Andel"], errors="coerce")
    fondertabell["AndelP"] = pd.to_numeric(fondertabell["AndelP"], errors="coerce")
    portfolio_metadata["Initial_Index_Value"] = pd.to_numeric(
        portfolio_metadata["Initial_Index_Value"], errors="coerce"
    )

    return {
        "transactions": transactions,
        "mapping": mapping,
        "portfolio_metadata": portfolio_metadata,
        "benchmarks": benchmarks,
        "fondertabell": fondertabell,
    }
