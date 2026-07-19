"""Separate BI prep step that builds a first Power BI data contract from the shared output workbook."""

from __future__ import annotations

import argparse
import logging
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from . import config
from .bi_io import extract_run_parameters, load_portfolio_output
from .bi_metrics import PERIOD_ORDER, compute_kpis, has_minimum_observations, slice_period

ANALYSIS_PREFIXES = ("PORT_", "BM_", "POLICY_")
ALLOCATION_SNAPSHOT_SHEET_NAME = "Fact_Portfolio_Alloc_Snapshot"
ALLOCATION_MONTHLY_SHEET_NAME = "Fact_Portfolio_Alloc_Monthly"
ALLOCATION_MONTHLY_COLUMNS = [
    "Portfolio_Key",
    "Series_ID",
    "Instrument_Key",
    "ISIN",
    "Display_Name",
    "Price_Currency",
    "Category",
    "Period_End_Date",
    "Market_Value_SEK",
    "Portfolio_MV_SEK",
    "Weight",
    "Weight_Source",
]
COURTAGE_SHEET_NAME = "Fact_Portfolio_Courtage"
COURTAGE_FACT_COLUMNS = [
    "Portfolio_Key",
    "Portfolio_Name",
    "Series_ID",
    "Instrument_Key",
    "ISIN",
    "Display_Name",
    "Category",
    "Period_End_Date",
    "Currency",
    "Courtage_Native",
    "Courtage_SEK",
    "Txn_Count",
]
TABLE_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")


def _configure_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input-path",
        default=None,
        help="Path to portfolio_output_timeseries.xlsx",
    )
    parser.add_argument(
        "--output-path",
        default=None,
        help="Path to portfolio_bi_data.xlsx",
    )
    return parser.parse_args()


def _clean_text(series: pd.Series) -> pd.Series:
    cleaned = series.fillna("").astype(str).str.strip()
    return cleaned.replace({"nan": "", "None": "", "NaT": "", "<NA>": ""})


def _nullable_text(series: pd.Series) -> pd.Series:
    cleaned = _clean_text(series)
    return cleaned.mask(cleaned == "")


def _combine_optional_columns(df: pd.DataFrame, target: str) -> pd.Series:
    map_column = f"{target}_map"
    series_column = f"{target}_series"
    if map_column not in df.columns and series_column not in df.columns:
        # Kolumnen finns bara i en av de två merge-källorna och fick inget suffix.
        # Returnera den direkt om den existerar, annars en tom serie.
        return df[target] if target in df.columns else pd.Series(pd.NA, index=df.index, dtype="object")
    map_values = df[map_column] if map_column in df.columns else pd.Series(pd.NA, index=df.index, dtype="object")
    series_values = (
        df[series_column] if series_column in df.columns else pd.Series(pd.NA, index=df.index, dtype="object")
    )
    return map_values.combine_first(series_values)


def _build_analysis_metadata(
    series_definition: pd.DataFrame,
    master_long: pd.DataFrame,
) -> pd.DataFrame:
    analysis_ids = sorted(
        {
            series_id
            for series_id in _clean_text(master_long["Series_ID"]).tolist()
            if any(series_id.startswith(prefix) for prefix in ANALYSIS_PREFIXES)
        }
    )
    if not analysis_ids:
        raise ValueError("No BI analysis series found in Master_TimeSeries_Long")

    metadata = series_definition.copy()
    metadata["Series_ID"] = _clean_text(metadata["Series_ID"])
    metadata["Series_Type"] = _clean_text(metadata["Series_Type"]).str.upper()
    metadata["Portfolio_Name"] = _nullable_text(metadata["Portfolio_Name"])
    metadata["Variant"] = _nullable_text(metadata["Variant"])
    metadata["Benchmark_ID"] = _nullable_text(metadata["Benchmark_ID"])
    metadata["Yahoo_Ticker"] = _nullable_text(metadata["Yahoo_Ticker"])
    metadata["ISIN"] = _nullable_text(metadata["ISIN"])
    metadata["Display_Name"] = _nullable_text(metadata["Display_Name"])
    metadata["Price_Currency"] = _nullable_text(metadata["Price_Currency"])
    metadata["Instrument_Type"] = _nullable_text(metadata["Instrument_Type"])
    metadata["Category"] = _nullable_text(metadata["Category"])
    metadata["Geography"] = _nullable_text(metadata["Geography"])
    metadata = metadata[metadata["Series_ID"].isin(analysis_ids)].copy()
    if metadata.empty:
        raise ValueError("Series_Definition does not contain BI metadata for analysis series")

    missing_ids = sorted(set(analysis_ids) - set(metadata["Series_ID"].tolist()))
    if missing_ids:
        raise ValueError(f"Series_Definition is missing BI metadata for series: {missing_ids}")

    metadata["Is_Main_Portfolio_Series"] = (
        (metadata["Series_Type"] == "PORT") & metadata["Category"].isna()
    )
    metadata["Is_Category_Series"] = (
        (metadata["Series_Type"] == "PORT") & metadata["Category"].notna()
    )
    metadata["Is_Benchmark"] = metadata["Series_Type"] == "BM"
    # Policyreferenser (passiva tvåbucketsindex) är jämförelseserier på samma
    # sätt som benchmarks: med i översikt och KPI-beräkning.
    metadata["Is_Policy"] = metadata["Series_Type"] == "POLICY"
    metadata["Is_Overview_Eligible"] = (
        metadata["Is_Main_Portfolio_Series"] | metadata["Is_Benchmark"] | metadata["Is_Policy"]
    )
    metadata["Is_Performance_Eligible"] = (
        metadata["Is_Overview_Eligible"] | metadata["Is_Category_Series"]
    )

    return metadata.sort_values(["Series_Type", "Portfolio_Name", "Variant", "Benchmark_ID", "Series_ID"]).reset_index(
        drop=True
    )


def _build_dim_portfolio(analysis_metadata: pd.DataFrame) -> pd.DataFrame:
    portfolios = analysis_metadata.loc[
        analysis_metadata["Portfolio_Name"].notna(),
        ["Portfolio_Name", "Index_Start_Date", "Initial_Index_Value"],
    ].copy()
    if portfolios.empty:
        return pd.DataFrame(
            columns=["Portfolio_Key", "Portfolio_Name", "Index_Start_Date", "Initial_Index_Value"]
        )
    portfolios = portfolios.sort_values(["Portfolio_Name", "Index_Start_Date"]).drop_duplicates(
        subset=["Portfolio_Name"],
        keep="first",
    )
    portfolios.insert(0, "Portfolio_Key", portfolios["Portfolio_Name"])
    return portfolios.reset_index(drop=True)


def _build_dim_series(analysis_metadata: pd.DataFrame) -> pd.DataFrame:
    dim_series = analysis_metadata[
        [
            "Series_ID",
            "Series_Type",
            "Portfolio_Name",
            "Variant",
            "Benchmark_ID",
            "Category",
            "Geography",
            "Yahoo_Ticker",
            "ISIN",
            "Display_Name",
            "Price_Currency",
            "Instrument_Type",
            "Include_From_Date",
            "Index_Start_Date",
            "Initial_Index_Value",
            "Is_Main_Portfolio_Series",
            "Is_Category_Series",
            "Is_Benchmark",
            "Is_Policy",
            "Is_Overview_Eligible",
            "Is_Performance_Eligible",
        ]
    ].copy()
    dim_series.insert(2, "Portfolio_Key", dim_series["Portfolio_Name"])
    if dim_series["Series_ID"].duplicated().any():
        duplicates = dim_series.loc[dim_series["Series_ID"].duplicated(), "Series_ID"].tolist()
        raise ValueError(f"Dim_Series would contain duplicate Series_ID values: {duplicates}")
    return dim_series.reset_index(drop=True)


def _build_fact_series_daily(master_long: pd.DataFrame, dim_series: pd.DataFrame) -> pd.DataFrame:
    valid_ids = set(dim_series["Series_ID"].tolist())
    fact_daily = master_long.copy()
    fact_daily["Series_ID"] = _clean_text(fact_daily["Series_ID"])
    fact_daily = fact_daily[fact_daily["Series_ID"].isin(valid_ids)].copy()
    if fact_daily.empty:
        raise ValueError("Master_TimeSeries_Long has no rows for the BI analysis universe")
    fact_daily["Date"] = pd.to_datetime(fact_daily["Date"], errors="coerce")
    for column in ("RET", "IDX", "DD"):
        fact_daily[column] = pd.to_numeric(fact_daily[column], errors="coerce")
    return fact_daily[["Date", "Series_ID", "RET", "IDX", "DD"]].sort_values(
        ["Series_ID", "Date"]
    ).reset_index(drop=True)


def _build_dim_date(fact_series_daily: pd.DataFrame) -> pd.DataFrame:
    dates = pd.Index(pd.to_datetime(fact_series_daily["Date"], errors="coerce").dropna().unique()).sort_values()
    if len(dates) == 0:
        return pd.DataFrame(
            columns=["Date", "Year", "Month", "Month_Name", "Quarter", "YearMonth", "Is_YTD_Latest_Flag"]
        )
    latest_date = dates.max()
    dim_date = pd.DataFrame({"Date": dates})
    dim_date["Year"] = dim_date["Date"].dt.year
    dim_date["Month"] = dim_date["Date"].dt.month
    dim_date["Month_Name"] = dim_date["Date"].dt.strftime("%B")
    dim_date["Quarter"] = "Q" + dim_date["Date"].dt.quarter.astype(str)
    dim_date["YearMonth"] = dim_date["Date"].dt.strftime("%Y-%m")
    dim_date["Is_YTD_Latest_Flag"] = (
        (dim_date["Date"].dt.year == latest_date.year) & (dim_date["Date"] <= latest_date)
    )
    return dim_date.reset_index(drop=True)


def _build_fact_series_kpi(
    fact_series_daily: pd.DataFrame,
    dim_series: pd.DataFrame,
    rf_rate_annual: float,
    trading_days_per_year: int,
) -> pd.DataFrame:
    eligible_ids = set(
        dim_series.loc[dim_series["Is_Performance_Eligible"], "Series_ID"].tolist()
    )
    rows: list[dict[str, object]] = []
    for series_id, series_frame in fact_series_daily.groupby("Series_ID", sort=True):
        if series_id not in eligible_ids:
            continue
        for period in PERIOD_ORDER:
            period_slice = slice_period(series_frame, period)
            if not has_minimum_observations(period_slice.frame, period):
                continue
            rows.append(
                {
                    "Series_ID": series_id,
                    "Period": period,
                    **compute_kpis(
                        period_slice.frame,
                        rf_rate_annual,
                        trading_days_per_year,
                        anchor_idx=period_slice.anchor_idx,
                        anchor_date=period_slice.anchor_date,
                    ),
                }
            )

    columns = [
        "Series_ID",
        "Period",
        "Start_Date",
        "End_Date",
        "Obs_Days",
        "Return_Total",
        "CAGR",
        "Vol",
        "Sharpe",
        "Sortino",
        "Max_DD",
        "Calmar",
        "DD_Duration_Max_Days",
        "Positive_Days_Pct",
    ]
    if not rows:
        return pd.DataFrame(columns=columns)

    fact_kpi = pd.DataFrame(rows)
    fact_kpi["Period"] = pd.Categorical(fact_kpi["Period"], categories=PERIOD_ORDER, ordered=True)
    return fact_kpi.sort_values(["Series_ID", "Period"]).reset_index(drop=True)[columns]


def _attach_instrument_ter(
    dim_instrument: pd.DataFrame,
    instrument_cost: pd.DataFrame | None,
) -> pd.DataFrame:
    """Join löpande avgift (TER) på Dim_Instrument via ISIN.

    TER bärs genom upstream-outputens ``Instrument_Cost``-sheet. Instrument utan
    TER-rad (eller där Nordnet inte visade avgiften) får TER=NA och en talande
    status, aldrig ett påhittat värde. Bakåtkompatibelt: saknas sheeten helt blir
    kolumnerna tomma.
    """
    out = dim_instrument.copy()
    if instrument_cost is None or instrument_cost.empty or "ISIN" not in instrument_cost.columns:
        out["TER"] = pd.NA
        out["TER_Status"] = pd.NA
        out["TER_Source"] = pd.NA
        return out

    cost = instrument_cost.copy()
    cost["ISIN"] = cost["ISIN"].astype(str).str.strip()
    cost = cost[cost["ISIN"] != ""].drop_duplicates(subset=["ISIN"], keep="first")
    cols = [c for c in ("ISIN", "TER", "TER_Status", "TER_Source") if c in cost.columns]

    out["_join_isin"] = out["ISIN"].astype(str).str.strip()
    merged = out.merge(
        cost[cols].rename(columns={"ISIN": "_join_isin"}),
        on="_join_isin",
        how="left",
    ).drop(columns=["_join_isin"])
    merged["TER"] = pd.to_numeric(merged["TER"], errors="coerce") if "TER" in merged.columns else pd.NA
    for column in ("TER_Status", "TER_Source"):
        if column not in merged.columns:
            merged[column] = pd.NA
    return merged


TER_SEED_REQUIRED_COLUMNS = ("ISIN", "TER")


def _load_ter_seed(seed_path: Path) -> pd.DataFrame | None:
    """Läs den statiska TER-seedfilen för utträdda/otäckta instrument ([AU]).

    Saknad eller oläsbar fil är aldrig ett hårt fel: bi_prep fortsätter utan
    seed och loggar en varning (broad except är avsiktligt här).
    """
    if not seed_path.exists():
        logging.warning("TER-seedfil saknas (%s) – fortsätter utan seed", seed_path)
        return None
    try:
        seed = pd.read_csv(seed_path, sep=";", encoding="utf-8", dtype=str)
    except Exception as exc:
        logging.warning("TER-seedfil kunde inte läsas (%s): %s – fortsätter utan seed", seed_path, exc)
        return None

    missing_columns = [c for c in TER_SEED_REQUIRED_COLUMNS if c not in seed.columns]
    if missing_columns:
        logging.warning(
            "TER-seedfil saknar obligatoriska kolumner %s (%s) – fortsätter utan seed",
            missing_columns,
            seed_path,
        )
        return None

    seed["ISIN"] = seed["ISIN"].astype(str).str.strip()
    seed["TER"] = pd.to_numeric(seed["TER"], errors="coerce")
    seed = seed[(seed["ISIN"] != "") & seed["TER"].notna()]
    return seed.drop_duplicates(subset=["ISIN"], keep="first")


def _apply_ter_seed(dim_instrument: pd.DataFrame, seed_path: Path) -> pd.DataFrame:
    """Fyll TER för utträdda/otäckta instrument från den statiska seedfilen.

    Matchning på ISIN. Seed fyller **endast** rader utan ett skrapat värde
    (TER_Status='no_data' eller saknad status/TER) – ett skrapat värde skrivs
    aldrig över, samma invariant som Mapping-fälten i transaction_data.
    Instrument utan ISIN (benchmarks, policyserier) matchar aldrig och
    berörs inte. Seedade rader får TER_Status='ok', TER_Source='seed'.
    """
    out = dim_instrument.copy()
    if out.empty or "ISIN" not in out.columns:
        return out

    seed = _load_ter_seed(seed_path)
    if seed is None or seed.empty:
        return out

    isin_clean = out["ISIN"].astype(str).str.strip()
    known_isins = set(isin_clean[out["ISIN"].notna()])
    unknown_isins = sorted(set(seed["ISIN"]) - known_isins)
    for isin in unknown_isins:
        logging.warning(
            "TER-seed: ISIN %s i seedfilen finns inte i instrumentuniversumet – rad ignoreras",
            isin,
        )

    ter_status = out["TER_Status"] if "TER_Status" in out.columns else pd.Series(pd.NA, index=out.index)
    needs_seed = ter_status.isna() | (ter_status.astype(str).str.strip() == "no_data")

    seed_ter = seed.set_index("ISIN")["TER"]
    seed_rows = needs_seed & out["ISIN"].notna() & isin_clean.isin(seed_ter.index)
    if not seed_rows.any():
        return out

    out.loc[seed_rows, "TER"] = isin_clean[seed_rows].map(seed_ter)
    out.loc[seed_rows, "TER_Status"] = "ok"
    out.loc[seed_rows, "TER_Source"] = "seed"
    logging.info(
        "TER-seed: fyllde %s instrument från %s",
        int(seed_rows.sum()),
        seed_path.name,
    )
    return out


def _build_dim_instrument(
    series_definition: pd.DataFrame,
    portfolio_series_map: pd.DataFrame,
    instrument_cost: pd.DataFrame | None = None,
) -> pd.DataFrame:
    map_rows = portfolio_series_map[
        ["Yahoo_Ticker", "ISIN", "Display_Name", "Price_Currency"]
    ].copy()
    map_rows["Yahoo_Ticker"] = _nullable_text(map_rows["Yahoo_Ticker"])
    map_rows["ISIN"] = _nullable_text(map_rows["ISIN"])
    map_rows["Display_Name"] = _nullable_text(map_rows["Display_Name"])
    map_rows["Price_Currency"] = _nullable_text(map_rows["Price_Currency"])

    # Driver ([BD]) saknas i uppström-workbooks byggda före denna funktion fanns -
    # grace:a till tom kolumn i stället för att krascha, samma mönster som
    # Geography redan hanteras (ovaliderad i bi_io.REQUIRED_SHEETS).
    series_definition = series_definition.copy()
    if "Driver" not in series_definition.columns:
        logging.warning(
            "Series_Definition saknar kolumnen 'Driver' - uppström-workbooken är "
            "byggd före [BD]; Dim_Instrument.Driver blir tomt för samtliga instrument"
        )
        series_definition["Driver"] = pd.NA

    series_rows = series_definition[
        [
            "Yahoo_Ticker",
            "ISIN",
            "Display_Name",
            "Price_Currency",
            "Instrument_Type",
            "Category",
            "Geography",
            "Driver",
        ]
    ].copy()
    series_rows["Yahoo_Ticker"] = _nullable_text(series_rows["Yahoo_Ticker"])
    series_rows["ISIN"] = _nullable_text(series_rows["ISIN"])
    series_rows["Display_Name"] = _nullable_text(series_rows["Display_Name"])
    series_rows["Price_Currency"] = _nullable_text(series_rows["Price_Currency"])
    series_rows["Instrument_Type"] = _nullable_text(series_rows["Instrument_Type"])
    series_rows["Category"] = _nullable_text(series_rows["Category"])
    series_rows["Geography"] = _nullable_text(series_rows["Geography"])
    series_rows["Driver"] = _nullable_text(series_rows["Driver"])

    all_tickers = (
        pd.concat([map_rows[["Yahoo_Ticker"]], series_rows[["Yahoo_Ticker"]]], ignore_index=True)
        .dropna()
        .drop_duplicates()
        .sort_values("Yahoo_Ticker")
        .reset_index(drop=True)
    )
    if all_tickers.empty:
        return pd.DataFrame(
            columns=[
                "Instrument_Key",
                "Yahoo_Ticker",
                "ISIN",
                "Display_Name",
                "Price_Currency",
                "Instrument_Type",
                "Category",
                "Geography",
                "Driver",
                "Structure",
                "TER",
                "TER_Status",
                "TER_Source",
            ]
        )

    metadata_from_map = (
        map_rows.dropna(subset=["Yahoo_Ticker"])
        .sort_values(["Yahoo_Ticker", "Display_Name", "ISIN", "Price_Currency"])
        .drop_duplicates(subset=["Yahoo_Ticker"], keep="first")
    )
    metadata_from_series = (
        series_rows.dropna(subset=["Yahoo_Ticker"])
        .sort_values(
            ["Yahoo_Ticker", "Display_Name", "ISIN", "Price_Currency", "Instrument_Type", "Category", "Geography", "Driver"]
        )
        .drop_duplicates(subset=["Yahoo_Ticker"], keep="first")
    )
    dim_instrument = all_tickers.merge(metadata_from_map, on="Yahoo_Ticker", how="left")
    dim_instrument = dim_instrument.merge(
        metadata_from_series,
        on="Yahoo_Ticker",
        how="left",
        suffixes=("_map", "_series"),
    )
    dim_instrument.insert(0, "Instrument_Key", dim_instrument["Yahoo_Ticker"])
    for column in ("ISIN", "Display_Name", "Price_Currency", "Instrument_Type", "Category", "Geography", "Driver"):
        dim_instrument[column] = _combine_optional_columns(dim_instrument, column)
    dim_instrument["Structure"] = pd.NA
    dim_instrument = _attach_instrument_ter(dim_instrument, instrument_cost)
    return dim_instrument[
        [
            "Instrument_Key",
            "Yahoo_Ticker",
            "ISIN",
            "Display_Name",
            "Price_Currency",
            "Instrument_Type",
            "Category",
            "Geography",
            "Driver",
            "Structure",
            "TER",
            "TER_Status",
            "TER_Source",
        ]
    ]


def _build_fact_portfolio_allocation_snapshot(
    portfolio_series_map: pd.DataFrame,
    dim_series: pd.DataFrame,
    snapshot_date: pd.Timestamp,
) -> pd.DataFrame:
    valid_series = set(dim_series["Series_ID"].tolist())
    snapshot = portfolio_series_map.copy()
    snapshot["Portfolio_Name"] = _nullable_text(snapshot["Portfolio_Name"])
    snapshot["Series_ID"] = _clean_text(snapshot["Series_ID"])
    snapshot["ISIN"] = _nullable_text(snapshot["ISIN"])
    snapshot["Display_Name"] = _nullable_text(snapshot["Display_Name"])
    snapshot["Price_Currency"] = _nullable_text(snapshot["Price_Currency"])
    snapshot["Yahoo_Ticker"] = _nullable_text(snapshot["Yahoo_Ticker"])
    snapshot["Weight_Source"] = _nullable_text(snapshot["Weight_Source"])
    snapshot["Weight"] = pd.to_numeric(snapshot["Weight"], errors="coerce")
    snapshot = snapshot[
        snapshot["Series_ID"].isin(valid_series) & snapshot["Yahoo_Ticker"].notna() & snapshot["Weight"].notna()
    ].copy()
    snapshot.insert(0, "Portfolio_Key", snapshot["Portfolio_Name"])
    snapshot.insert(3, "Instrument_Key", snapshot["Yahoo_Ticker"])
    snapshot["Snapshot_Date"] = pd.Timestamp(snapshot_date).normalize()
    return snapshot[
        [
            "Portfolio_Key",
            "Series_ID",
            "Instrument_Key",
            "ISIN",
            "Display_Name",
            "Price_Currency",
            "Weight",
            "Weight_Source",
            "Snapshot_Date",
        ]
    ].sort_values(["Portfolio_Key", "Series_ID", "Instrument_Key"]).reset_index(drop=True)


def _build_fact_portfolio_alloc_monthly(
    portfolio_alloc_monthly: pd.DataFrame,
    dim_series: pd.DataFrame,
) -> pd.DataFrame:
    """Historical month-end REAL allocation weights, fund grain, star-schema keyed.

    Category weights are an exact roll-up in Power BI: sum ``Weight`` over the
    ``Category`` attribute within each Portfolio_Key/Period_End_Date. Links to
    Dim_Portfolio (Portfolio_Key), Dim_Instrument (Instrument_Key), Dim_Series
    (Series_ID) and Dim_Date (Period_End_Date).
    """
    if portfolio_alloc_monthly.empty:
        return pd.DataFrame(columns=ALLOCATION_MONTHLY_COLUMNS)

    valid_series = set(dim_series["Series_ID"].tolist())
    monthly = portfolio_alloc_monthly.copy()
    monthly["Portfolio_Name"] = _nullable_text(monthly["Portfolio_Name"])
    monthly["Series_ID"] = _clean_text(monthly["Series_ID"])
    monthly["Yahoo_Ticker"] = _nullable_text(monthly["Yahoo_Ticker"])
    monthly["ISIN"] = _nullable_text(monthly["ISIN"])
    monthly["Display_Name"] = _nullable_text(monthly["Display_Name"])
    monthly["Price_Currency"] = _nullable_text(monthly["Price_Currency"])
    monthly["Category"] = _nullable_text(monthly["Category"])
    monthly["Weight_Source"] = (
        _nullable_text(monthly["Weight_Source"])
        if "Weight_Source" in monthly.columns
        else pd.Series(pd.NA, index=monthly.index, dtype="object")
    )
    monthly["Period_End_Date"] = pd.to_datetime(monthly["Period_End_Date"], errors="coerce").dt.normalize()
    for column in ("Market_Value_SEK", "Portfolio_MV_SEK", "Weight"):
        monthly[column] = (
            pd.to_numeric(monthly[column], errors="coerce")
            if column in monthly.columns
            else pd.Series(pd.NA, index=monthly.index, dtype="float")
        )

    monthly = monthly[
        monthly["Series_ID"].isin(valid_series)
        & monthly["Yahoo_Ticker"].notna()
        & monthly["Weight"].notna()
        & monthly["Period_End_Date"].notna()
    ].copy()
    monthly["Portfolio_Key"] = monthly["Portfolio_Name"]
    monthly["Instrument_Key"] = monthly["Yahoo_Ticker"]
    return (
        monthly[ALLOCATION_MONTHLY_COLUMNS]
        .sort_values(["Portfolio_Key", "Period_End_Date", "Series_ID", "Instrument_Key"])
        .reset_index(drop=True)
    )


def _build_fact_portfolio_courtage(portfolio_courtage: pd.DataFrame) -> pd.DataFrame:
    """Realiserat courtage per portfölj × månad × instrument × valuta, star-keyat.

    Länkar till Dim_Portfolio (Portfolio_Key), Dim_Instrument (Instrument_Key),
    Dim_Series (Series_ID) och Dim_Date (Period_End_Date). ``Courtage_SEK`` är den
    rollup-säkra basvaluta-summan (redan inbakad i REAL); ``Courtage_Native`` är
    råvärdet per valuta för avstämning mot Nordnet-exportens Courtage-kolumn.
    """
    if portfolio_courtage.empty:
        return pd.DataFrame(columns=COURTAGE_FACT_COLUMNS)

    df = portfolio_courtage.copy()
    df["Portfolio_Name"] = _nullable_text(df["Portfolio_Name"])
    df["Series_ID"] = _nullable_text(df["Series_ID"]) if "Series_ID" in df.columns else pd.NA
    df["ISIN"] = _nullable_text(df["ISIN"])
    df["Yahoo_Ticker"] = _nullable_text(df["Yahoo_Ticker"]) if "Yahoo_Ticker" in df.columns else pd.NA
    df["Display_Name"] = _nullable_text(df["Display_Name"]) if "Display_Name" in df.columns else pd.NA
    df["Category"] = _nullable_text(df["Category"]) if "Category" in df.columns else pd.NA
    df["Currency"] = _nullable_text(df["Currency"]) if "Currency" in df.columns else pd.NA
    df["Period_End_Date"] = pd.to_datetime(df["Period_End_Date"], errors="coerce").dt.normalize()
    for column in ("Courtage_Native", "Courtage_SEK", "Txn_Count"):
        df[column] = (
            pd.to_numeric(df[column], errors="coerce")
            if column in df.columns
            else pd.Series(pd.NA, index=df.index, dtype="float")
        )

    df = df[df["Portfolio_Name"].notna() & df["Period_End_Date"].notna()].copy()
    if df.empty:
        return pd.DataFrame(columns=COURTAGE_FACT_COLUMNS)
    df["Portfolio_Key"] = df["Portfolio_Name"]
    df["Instrument_Key"] = df["Yahoo_Ticker"]
    return (
        df[COURTAGE_FACT_COLUMNS]
        .sort_values(["Portfolio_Key", "Period_End_Date", "ISIN", "Currency"])
        .reset_index(drop=True)
    )


def _add_excel_table(writer: pd.ExcelWriter, sheet_name: str, table_name: str) -> None:
    """Wrap a worksheet's used range in an Excel table for more stable Power BI navigation."""
    worksheet = writer.sheets[sheet_name]
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return

    table_ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    worksheet.add_table(table)

    # Keep header fill deterministic instead of relying only on the table style theme.
    for cell in worksheet[1]:
        cell.fill = TABLE_HEADER_FILL


def _warn_on_unclassified_active_holdings(
    dim_instrument: pd.DataFrame, fact_allocation_snapshot: pd.DataFrame
) -> None:
    """Kanariefågel ([BD]): innehav med vikt i aktuellt snapshot utan Drivkraft-klassning.

    Efter [BC] ska alla aktiva fonder vara klassade i Fondertabell (fonder.xlsx) -
    en aktiv, oklassad post signalerar att en ny/utbytt fond glömts i klassningen.
    """
    if dim_instrument.empty or fact_allocation_snapshot.empty or "Driver" not in dim_instrument.columns:
        return
    active = fact_allocation_snapshot[fact_allocation_snapshot["Weight"] > 0]
    active_keys = set(active["Instrument_Key"].dropna().unique())
    if not active_keys:
        return
    driver_by_key = dim_instrument.set_index("Instrument_Key")["Driver"]
    unclassified = sorted(
        key for key in active_keys if key not in driver_by_key.index or pd.isna(driver_by_key[key])
    )
    if unclassified:
        logging.warning(
            "Drivkraft ([BD]): %s aktivt innehav (vikt > 0 i snapshotet) saknar klassning i "
            "Fondertabell: %s",
            len(unclassified),
            ", ".join(unclassified),
        )


def run(
    source_output_path: str | Path | None = None,
    bi_output_path: str | Path | None = None,
) -> None:
    """Read the shared workbook and write a first BI workbook."""
    _configure_logging()

    source_path = Path(source_output_path or config.BI_DATA_SOURCE_PATH)
    output_path = Path(bi_output_path or config.BI_DATA_OUTPUT_PATH)

    logging.info("Loading BI source workbook: %s", source_path)
    source = load_portfolio_output(source_path)
    rf_rate_annual, trading_days_per_year = extract_run_parameters(source.run_config)

    analysis_metadata = _build_analysis_metadata(source.series_definition, source.master_long)
    dim_portfolio = _build_dim_portfolio(analysis_metadata)
    dim_series = _build_dim_series(analysis_metadata)
    fact_series_daily = _build_fact_series_daily(source.master_long, dim_series)
    dim_date = _build_dim_date(fact_series_daily)
    fact_series_kpi = _build_fact_series_kpi(
        fact_series_daily,
        dim_series,
        rf_rate_annual,
        trading_days_per_year,
    )
    dim_instrument = _build_dim_instrument(
        source.series_definition,
        source.portfolio_series_map,
        source.instrument_cost,
    )
    dim_instrument = _apply_ter_seed(dim_instrument, config.PATH_TER_SEED)
    fact_portfolio_courtage = _build_fact_portfolio_courtage(source.portfolio_courtage)
    snapshot_date = fact_series_daily["Date"].max()
    fact_allocation_snapshot = _build_fact_portfolio_allocation_snapshot(
        source.portfolio_series_map,
        dim_series,
        snapshot_date,
    )
    fact_allocation_monthly = _build_fact_portfolio_alloc_monthly(
        source.portfolio_alloc_monthly,
        dim_series,
    )
    _warn_on_unclassified_active_holdings(dim_instrument, fact_allocation_snapshot)

    logging.info(
        "BI analysis universe: %s series, %s daily rows, %s KPI rows, snapshot date %s",
        len(dim_series),
        len(fact_series_daily),
        len(fact_series_kpi),
        pd.Timestamp(snapshot_date).date().isoformat(),
    )
    if fact_allocation_monthly.empty:
        logging.info(
            "Fact_Portfolio_Alloc_Monthly is empty; upstream workbook has no %s sheet",
            "Portfolio_Alloc_Monthly",
        )
    else:
        logging.info(
            "Fact_Portfolio_Alloc_Monthly: %s rows across %s month-end period(s)",
            len(fact_allocation_monthly),
            fact_allocation_monthly["Period_End_Date"].nunique(),
        )
    if dim_instrument.empty:
        logging.info("Dim_Instrument is empty; upstream artifact did not materialize instrument tickers")
    else:
        ter_ok = int((dim_instrument["TER_Status"] == "ok").sum()) if "TER_Status" in dim_instrument.columns else 0
        logging.info(
            "Dim_Instrument built from upstream ticker metadata (%s instrument, %s med TER)",
            len(dim_instrument),
            ter_ok,
        )
    if fact_portfolio_courtage.empty:
        logging.info(
            "Fact_Portfolio_Courtage is empty; upstream workbook has no %s sheet",
            "Portfolio_Courtage",
        )
    else:
        logging.info(
            "Fact_Portfolio_Courtage: %s rader, %.2f SEK totalt realiserat courtage",
            len(fact_portfolio_courtage),
            float(fact_portfolio_courtage["Courtage_SEK"].sum()),
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        dim_date.to_excel(writer, sheet_name="Dim_Date", index=False)
        dim_portfolio.to_excel(writer, sheet_name="Dim_Portfolio", index=False)
        dim_series.to_excel(writer, sheet_name="Dim_Series", index=False)
        dim_instrument.to_excel(writer, sheet_name="Dim_Instrument", index=False)
        fact_series_daily.to_excel(writer, sheet_name="Fact_Series_Daily", index=False)
        fact_series_kpi.to_excel(writer, sheet_name="Fact_Series_KPI", index=False)
        fact_allocation_snapshot.to_excel(
            writer,
            sheet_name=ALLOCATION_SNAPSHOT_SHEET_NAME,
            index=False,
        )
        fact_allocation_monthly.to_excel(
            writer,
            sheet_name=ALLOCATION_MONTHLY_SHEET_NAME,
            index=False,
        )
        fact_portfolio_courtage.to_excel(
            writer,
            sheet_name=COURTAGE_SHEET_NAME,
            index=False,
        )
        _add_excel_table(writer, "Dim_Date", "Dim_Date")
        _add_excel_table(writer, "Dim_Portfolio", "Dim_Portfolio")
        _add_excel_table(writer, "Dim_Series", "Dim_Series")
        _add_excel_table(writer, "Dim_Instrument", "Dim_Instrument")
        _add_excel_table(writer, "Fact_Series_Daily", "Fact_Series_Daily")
        _add_excel_table(writer, "Fact_Series_KPI", "Fact_Series_KPI")
        _add_excel_table(
            writer,
            ALLOCATION_SNAPSHOT_SHEET_NAME,
            ALLOCATION_SNAPSHOT_SHEET_NAME,
        )
        _add_excel_table(
            writer,
            ALLOCATION_MONTHLY_SHEET_NAME,
            ALLOCATION_MONTHLY_SHEET_NAME,
        )
        _add_excel_table(writer, COURTAGE_SHEET_NAME, COURTAGE_SHEET_NAME)

    logging.info("BI data workbook written: %s", output_path)
    # Väg B: bi_prep skriver bara lokal BI-output. OneDrive-kopian sköts nattligt
    # av Fondanalys backup-jobb.


if __name__ == "__main__":
    args = _parse_args()
    run(
        source_output_path=args.input_path,
        bi_output_path=args.output_path,
    )
