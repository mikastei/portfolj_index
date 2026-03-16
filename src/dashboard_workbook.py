"""Build a separate Excel dashboard workbook from dashboard data output."""

from __future__ import annotations

import argparse
import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import absolute_coordinate
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from . import config

REQUIRED_SOURCE_SHEETS = [
    "KPI_Summary",
    "Period_Returns",
    "Chart_IDX_Wide",
    "Chart_DD_Wide",
    "Allocation_Snapshot",
    "Dashboard_Config",
    "Build_Info",
]
VARIANT_OPTIONS = ["REAL", "CUR", "TGT"]
PERIOD_OPTIONS = ["30D", "YTD", "1Y", "Since_Start"]
MAIN_CONTROL_NAME_MAP = {
    "Selected_Portfolio": "Control_Selected_Portfolio",
    "Selected_Variant": "Control_Selected_Variant",
    "Selected_Period": "Control_Selected_Period",
    "Selected_Compare_1": "Control_Selected_Compare_1",
    "Selected_Compare_2": "Control_Selected_Compare_2",
}
CATEGORY_CONTROL_NAME_MAP = {
    "Selected_Category_Portfolio": "Control_Selected_Category_Portfolio",
    "Selected_Category_Variant": "Control_Selected_Category_Variant",
    "Selected_Category_Period": "Control_Selected_Category_Period",
    "Selected_Category_Chart": "Control_Selected_Category_Chart",
    "Selected_Category_Benchmark": "Control_Selected_Category_Benchmark",
}
OVERVIEW_KPI_COLUMNS = [
    "Series_ID",
    "Display_Name",
    "Series_Type",
    "Portfolio_Name",
    "Variant",
    "Period",
    "Start_Date",
    "End_Date",
    "Obs_Days",
    "Return_Total",
    "CAGR",
    "Vol",
    "Sharpe",
    "Sortino",
    "Max_DD",
    "Calmar",
]
OVERVIEW_PERIOD_RETURN_COLUMNS = [
    "Series_ID",
    "Display_Name",
    "Series_Type",
    "Portfolio_Name",
    "Variant",
    "30D",
    "YTD",
    "1Y",
    "Since_Start",
]
CHART_WIDE_REQUIRED_BASE_COLUMNS = ["Date"]
OVERVIEW_ROLE_ROWS = {
    "Primary": 6,
    "Compare_1": 7,
    "Compare_2": 8,
}
PERFORMANCE_ROLE_ROWS = {
    "Primary": 18,
    "Compare_1": 19,
    "Compare_2": 20,
}
ALLOCATION_SNAPSHOT_COLUMNS = [
    "Portfolio_Name",
    "Series_ID",
    "Variant",
    "Display_Name",
    "Yahoo_Ticker",
    "Weight",
    "Weight_Source",
]
STRUCTURE_TABLE_MAX_ROWS = 25
STRUCTURE_TOP_COUNT = 10
CATEGORY_FILTER_MAX_ROWS = 50
CATEGORY_VISIBLE_MAX_ROWS = 8
CALC_MAIN_STRUCTURE_TABLE_FIRST_ROW = 18
CALC_MAIN_STRUCTURE_TOP_FIRST_ROW = 46
MAIN_CONTROL_SOURCE_CELLS = {
    "Control_Selected_Portfolio": "A5",
    "Control_Selected_Variant": "B5",
    "Control_Selected_Period": "C5",
    "Control_Selected_Compare_1": "D5",
    "Control_Selected_Compare_2": "E5",
}
CATEGORY_CONTROL_SOURCE_CELLS = {
    "Control_Selected_Category_Portfolio": "A5",
    "Control_Selected_Category_Variant": "B5",
    "Control_Selected_Category_Period": "C5",
    "Control_Selected_Category_Chart": "D5",
    "Control_Selected_Category_Benchmark": "E5",
}


def _chart_series_columns(df: pd.DataFrame, sheet_name: str) -> list[str]:
    _require_columns(df, sheet_name, CHART_WIDE_REQUIRED_BASE_COLUMNS)
    series_columns = [column for column in df.columns if str(column).strip() != "Date"]
    if not series_columns:
        raise ValueError(f"Sheet '{sheet_name}' does not contain any series columns.")
    return series_columns


def _configure_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input-path",
        default=None,
        help="Path to portfolio_dashboard_data.xlsx",
    )
    parser.add_argument(
        "--output-path",
        default=None,
        help="Path to portfolio_dashboard.xlsx",
    )
    return parser.parse_args()


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(col).strip() for col in out.columns]
    return out


def _require_columns(df: pd.DataFrame, sheet_name: str, required_columns: list[str]) -> None:
    missing_columns = [column for column in required_columns if column not in df.columns]
    if missing_columns:
        raise ValueError(f"Sheet '{sheet_name}' is missing required columns: {missing_columns}")


def _load_dashboard_data(path: Path) -> tuple[dict[str, pd.DataFrame], list[str]]:
    if not path.exists():
        raise FileNotFoundError(f"Dashboard data workbook does not exist: {path}")
    sheets = pd.read_excel(path, sheet_name=None)
    normalized = {name: _normalize_columns(df) for name, df in sheets.items()}
    found_sheets = list(normalized)
    missing_sheets = [sheet for sheet in REQUIRED_SOURCE_SHEETS if sheet not in normalized]
    if missing_sheets:
        raise ValueError(f"Dashboard data workbook is missing required sheets: {missing_sheets}")
    return normalized, found_sheets


def _variant_display(variant: str) -> str:
    value = str(variant or "").strip().upper()
    return {"REAL": "Real", "CUR": "Current", "TGT": "Target"}.get(value, value.title())


def _is_category_row(row: pd.Series) -> bool:
    if str(row.get("Series_Type") or "").strip().upper() != "PORT":
        return False
    portfolio_name = str(row.get("Portfolio_Name") or "").strip()
    variant = str(row.get("Variant") or "").strip().upper()
    display_name = str(row.get("Display_Name") or "").strip()
    expected_display = f"{portfolio_name} {_variant_display(variant)}".strip()
    return bool(display_name) and display_name != expected_display


def _build_series_catalog(kpi_summary: pd.DataFrame) -> pd.DataFrame:
    required_columns = ["Series_ID", "Display_Name", "Series_Type", "Portfolio_Name", "Variant"]
    _require_columns(kpi_summary, "KPI_Summary", required_columns)

    catalog = (
        kpi_summary.loc[:, required_columns]
        .drop_duplicates(subset=["Series_ID"])
        .sort_values(["Series_Type", "Display_Name", "Series_ID"])
        .reset_index(drop=True)
    )
    catalog["Portfolio_Name"] = catalog["Portfolio_Name"].fillna("").astype(str).str.strip()
    catalog["Variant"] = catalog["Variant"].fillna("").astype(str).str.strip().str.upper()
    catalog["Display_Name"] = catalog["Display_Name"].fillna("").astype(str).str.strip()
    catalog["Is_Category"] = catalog.apply(_is_category_row, axis=1)
    catalog["Is_Main_Portfolio"] = (catalog["Series_Type"] == "PORT") & (~catalog["Is_Category"])
    catalog["Is_Compare_Eligible"] = catalog["Series_Type"].isin(["PORT", "BM"]) & (~catalog["Is_Category"])
    return catalog


def _config_map(dashboard_config: pd.DataFrame) -> dict[str, str]:
    required_columns = ["Config_Key", "Config_Value"]
    _require_columns(dashboard_config, "Dashboard_Config", required_columns)
    config_rows = dashboard_config.loc[:, required_columns].dropna(subset=["Config_Key"])
    return {
        str(row["Config_Key"]).strip(): str(row["Config_Value"]).strip()
        for _, row in config_rows.iterrows()
    }


def _choose_default_portfolio(main_portfolios: list[str], dashboard_config_map: dict[str, str]) -> str:
    explicit_default = dashboard_config_map.get("default_portfolio", "").strip()
    if explicit_default and explicit_default in main_portfolios:
        return explicit_default
    return main_portfolios[0] if main_portfolios else ""


def _choose_default_period(dashboard_config_map: dict[str, str]) -> str:
    configured = dashboard_config_map.get("default_period", "").strip()
    if configured in PERIOD_OPTIONS:
        return configured
    return "1Y"


def _choose_default_category_selection(category_rows: list[list[object]], portfolio_name: str, slot: int) -> str:
    if slot < 1:
        return ""
    matching = [str(row[1]).strip() for row in category_rows if str(row[0]).strip() == portfolio_name and str(row[1]).strip()]
    unique_matching = list(dict.fromkeys(matching))
    index = slot - 1
    return unique_matching[index] if index < len(unique_matching) else ""


def _choose_default_benchmark(benchmark_rows: list[list[object]]) -> str:
    benchmark_names = [str(row[0]).strip() for row in benchmark_rows if str(row[0]).strip()]
    return benchmark_names[0] if benchmark_names else ""


def _set_title(ws, title: str, subtitle: str) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = subtitle
    ws["A4"] = "This sheet is a v1 workbook placeholder."
    ws["A5"] = "Layout, formulas and charts will be added in later implementation threads."
    ws.column_dimensions["A"].width = 72


def _write_table(ws, start_row: int, headers: list[str], rows: list[list[object]]) -> int:
    for column_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=column_index, value=header)
        cell.font = Font(bold=True)
    for row_offset, values in enumerate(rows, start=1):
        for column_index, value in enumerate(values, start=1):
            ws.cell(row=start_row + row_offset, column=column_index, value=value)
    return start_row + len(rows) + 1


def _write_dataframe(ws, df: pd.DataFrame) -> None:
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append([None if pd.isna(value) else value for value in row])


def _add_named_range(wb: Workbook, name: str, sheet_name: str, cell_ref: str) -> None:
    absolute_ref = absolute_coordinate(cell_ref)
    if ":" not in absolute_ref:
        absolute_ref = f"{absolute_ref}:{absolute_ref}"
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet_name}'!{absolute_ref}")


def _add_formula_named_range(wb: Workbook, name: str, formula: str) -> None:
    normalized_formula = formula[1:] if formula.startswith("=") else formula
    wb.defined_names[name] = DefinedName(name, attr_text=normalized_formula)


def _quote_sheet_name(sheet_name: str) -> str:
    return f"'{sheet_name}'"


def _absolute_range(column_letter: str, start_row: int, end_row: int) -> str:
    return f"${column_letter}${start_row}:${column_letter}${end_row}"


def _range_formula(sheet_name: str, column_letter: str, start_row: int, end_row: int) -> str:
    return f"={_quote_sheet_name(sheet_name)}!{_absolute_range(column_letter, start_row, end_row)}"


def _add_list_validation(ws, cell_ref: str, formula: str) -> None:
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(ws[cell_ref])


def _add_sheet_link(ws, cell_ref: str, target_sheet: str, target_cell: str, label: str) -> None:
    cell = ws[cell_ref]
    cell.value = label
    cell.hyperlink = f"#{target_sheet}!{target_cell}"
    cell.style = "Hyperlink"


def _build_control_sheet(
    wb: Workbook,
    defaults: dict[str, str],
    source_path: Path,
    output_path: Path,
    found_sheets: list[str],
) -> None:
    ws = wb.create_sheet("Control")
    ws["A1"] = "Dashboard Workbook Control Model"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Reference view only. Edit main controls on Overview and category controls on Category."
    _add_sheet_link(ws, "A3", "Overview", "A4", "Open Overview controls")
    _add_sheet_link(ws, "B3", "Category", "A4", "Open Category controls")
    ws["A4"] = "Main controls"
    ws["A4"].font = Font(bold=True)

    main_rows = [
        ["Selected_Portfolio", "=Control_Selected_Portfolio", "Mirrors visible input on Overview"],
        ["Selected_Variant", "=Control_Selected_Variant", "Mirrors visible input on Overview"],
        ["Selected_Period", "=Control_Selected_Period", "Mirrors visible input on Overview"],
        ["Selected_Compare_1", "=Control_Selected_Compare_1", "Mirrors visible input on Overview"],
        ["Selected_Compare_2", "=Control_Selected_Compare_2", "Mirrors visible input on Overview"],
    ]
    next_row = _write_table(ws, 5, ["Control_Key", "Value", "Note"], main_rows)

    ws.cell(row=next_row + 1, column=1, value="Category controls").font = Font(bold=True)
    category_rows = [
        ["Selected_Category_Portfolio", "=Control_Selected_Category_Portfolio", "Mirrors visible input on Category"],
        ["Selected_Category_Variant", "=Control_Selected_Category_Variant", "Mirrors visible input on Category"],
        ["Selected_Category_Period", "=Control_Selected_Category_Period", "Mirrors visible input on Category"],
        ["Selected_Category_Chart", "=Control_Selected_Category_Chart", "Chart category dropdown on Category"],
        ["Selected_Category_Benchmark", "=Control_Selected_Category_Benchmark", "Chart benchmark dropdown on Category"],
    ]
    next_row = _write_table(ws, next_row + 2, ["Control_Key", "Value", "Note"], category_rows)

    ws.cell(row=next_row + 1, column=1, value="Build metadata").font = Font(bold=True)
    metadata_rows = [
        ["Input_Path", str(source_path), "Dashboard data workbook"],
        ["Output_Path", str(output_path), "Separate dashboard workbook"],
        ["Source_Sheets_Found", ", ".join(found_sheets), "Validated source sheets"],
    ]
    _write_table(ws, next_row + 2, ["Metadata_Key", "Value", "Note"], metadata_rows)

    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 42


def _build_lists_sheet(
    wb: Workbook,
    main_portfolios: list[str],
    compare_series: list[str],
    category_portfolios: list[str],
    main_series_rows: list[list[object]],
    compare_lookup_rows: list[list[object]],
    category_rows: list[list[object]],
    benchmark_rows: list[list[object]],
) -> None:
    ws = wb.create_sheet("Lists")
    ws["A1"] = "Dashboard Workbook Lists"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Base lists for future dropdowns and formulas."

    row = 4
    variant_header_row = row
    row = _write_table(ws, row, ["Variant"], [[value] for value in VARIANT_OPTIONS])
    period_header_row = row + 1
    row = _write_table(ws, period_header_row, ["Period"], [[value] for value in PERIOD_OPTIONS])
    main_portfolio_header_row = row + 1
    row = _write_table(ws, main_portfolio_header_row, ["Main_Portfolio"], [[value] for value in main_portfolios])
    compare_header_row = row + 1
    row = _write_table(ws, compare_header_row, ["Compare_Series"], [[value] for value in compare_series])
    category_portfolio_header_row = row + 1
    row = _write_table(ws, category_portfolio_header_row, ["Category_Portfolio"], [[value] for value in category_portfolios])
    benchmark_header_row = row + 1
    row = _write_table(ws, benchmark_header_row, ["Category_Benchmark"], [[row[0]] for row in benchmark_rows])

    series_lookup_header_row = row + 1
    ws.cell(row=series_lookup_header_row - 1, column=1, value="Main series lookup").font = Font(bold=True)
    row = _write_table(
        ws,
        series_lookup_header_row,
        ["Portfolio_Name", "Variant", "Display_Name", "Series_ID", "Lookup_Key"],
        [row + [f"{row[0]}|{row[1]}"] for row in main_series_rows],
    )

    compare_lookup_header_row = row + 1
    ws.cell(row=compare_lookup_header_row - 1, column=1, value="Compare series lookup").font = Font(bold=True)
    row = _write_table(
        ws,
        compare_lookup_header_row,
        ["Display_Name", "Series_ID", "Series_Type", "Portfolio_Name", "Variant"],
        compare_lookup_rows,
    )

    category_lookup_rows = []
    portfolio_row_counts: dict[str, int] = {}
    for index, row_values in enumerate(category_rows, start=1):
        portfolio_name = str(row_values[0]).strip()
        portfolio_row_counts[portfolio_name] = portfolio_row_counts.get(portfolio_name, 0) + 1
        category_lookup_rows.append(
            row_values + [f"{row_values[0]}|{row_values[1]}", index, f"{portfolio_name}|{portfolio_row_counts[portfolio_name]}"]
        )

    ws.cell(row=row + 1, column=1, value="Category series by portfolio").font = Font(bold=True)
    category_lookup_header_row = row + 2
    row = _write_table(
        ws,
        category_lookup_header_row,
        ["Portfolio_Name", "Category_Display_Name", "Series_ID", "Variant", "Lookup_Key", "Row_Index", "Portfolio_Row_Key"],
        category_lookup_rows,
    )

    ws.cell(row=row + 1, column=1, value="Benchmark series").font = Font(bold=True)
    benchmark_lookup_header_row = row + 2
    _write_table(
        ws,
        benchmark_lookup_header_row,
        ["Display_Name", "Series_ID", "Row_Index"],
        [row + [index] for index, row in enumerate(benchmark_rows, start=1)],
    )

    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 42
    ws.column_dimensions["G"].width = 24

    _add_named_range(
        wb,
        "List_Variant",
        ws.title,
        _absolute_range("A", variant_header_row + 1, variant_header_row + len(VARIANT_OPTIONS)),
    )
    _add_named_range(
        wb,
        "List_Period",
        ws.title,
        _absolute_range("A", period_header_row + 1, period_header_row + len(PERIOD_OPTIONS)),
    )
    _add_named_range(
        wb,
        "List_Main_Portfolio",
        ws.title,
        _absolute_range("A", main_portfolio_header_row + 1, main_portfolio_header_row + max(len(main_portfolios), 1)),
    )
    _add_named_range(
        wb,
        "List_Compare_Series",
        ws.title,
        _absolute_range("A", compare_header_row + 1, compare_header_row + max(len(compare_series), 1)),
    )
    _add_named_range(
        wb,
        "List_Category_Portfolio",
        ws.title,
        _absolute_range("A", category_portfolio_header_row + 1, category_portfolio_header_row + max(len(category_portfolios), 1)),
    )
    _add_named_range(
        wb,
        "List_Category_Benchmark",
        ws.title,
        _absolute_range("A", benchmark_header_row + 1, benchmark_header_row + max(len(benchmark_rows), 1)),
    )
    _add_named_range(
        wb,
        "Lookup_Main_Series_Portfolio",
        ws.title,
        _absolute_range("A", series_lookup_header_row + 1, series_lookup_header_row + max(len(main_series_rows), 1)),
    )
    _add_named_range(
        wb,
        "Lookup_Main_Series_Variant",
        ws.title,
        _absolute_range("B", series_lookup_header_row + 1, series_lookup_header_row + max(len(main_series_rows), 1)),
    )
    _add_named_range(
        wb,
        "Lookup_Main_Series_Display_Name",
        ws.title,
        _absolute_range("C", series_lookup_header_row + 1, series_lookup_header_row + max(len(main_series_rows), 1)),
    )
    _add_named_range(
        wb,
        "Lookup_Main_Series_ID",
        ws.title,
        _absolute_range("D", series_lookup_header_row + 1, series_lookup_header_row + max(len(main_series_rows), 1)),
    )
    _add_named_range(
        wb,
        "Lookup_Main_Series_Key",
        ws.title,
        _absolute_range("E", series_lookup_header_row + 1, series_lookup_header_row + max(len(main_series_rows), 1)),
    )
    _add_named_range(
        wb,
        "Lookup_Compare_Display_Name",
        ws.title,
        _absolute_range("A", compare_lookup_header_row + 1, compare_lookup_header_row + max(len(compare_lookup_rows), 1)),
    )
    _add_named_range(
        wb,
        "Lookup_Compare_Series_ID",
        ws.title,
        _absolute_range("B", compare_lookup_header_row + 1, compare_lookup_header_row + max(len(compare_lookup_rows), 1)),
    )
    _add_named_range(
        wb,
        "Lookup_Category_Portfolio",
        ws.title,
        _absolute_range("A", category_lookup_header_row + 1, category_lookup_header_row + max(len(category_rows), 1)),
    )
    _add_named_range(
        wb,
        "Lookup_Category_Display_Name",
        ws.title,
        _absolute_range("B", category_lookup_header_row + 1, category_lookup_header_row + max(len(category_rows), 1)),
    )
    _add_named_range(
        wb,
        "Lookup_Category_Series_ID",
        ws.title,
        _absolute_range("C", category_lookup_header_row + 1, category_lookup_header_row + max(len(category_rows), 1)),
    )
    _add_named_range(
        wb,
        "Lookup_Category_Series_Variant",
        ws.title,
        _absolute_range("D", category_lookup_header_row + 1, category_lookup_header_row + max(len(category_rows), 1)),
    )
    _add_named_range(
        wb,
        "Lookup_Category_Key",
        ws.title,
        _absolute_range("E", category_lookup_header_row + 1, category_lookup_header_row + max(len(category_rows), 1)),
    )
    _add_named_range(
        wb,
        "Lookup_Category_Row_Index",
        ws.title,
        _absolute_range("F", category_lookup_header_row + 1, category_lookup_header_row + max(len(category_rows), 1)),
    )
    _add_named_range(
        wb,
        "Lookup_Category_Portfolio_Row_Key",
        ws.title,
        _absolute_range("G", category_lookup_header_row + 1, category_lookup_header_row + max(len(category_rows), 1)),
    )
    _add_named_range(
        wb,
        "Lookup_Benchmark_Display_Name",
        ws.title,
        _absolute_range("A", benchmark_lookup_header_row + 1, benchmark_lookup_header_row + max(len(benchmark_rows), 1)),
    )
    _add_named_range(
        wb,
        "Lookup_Benchmark_Series_ID",
        ws.title,
        _absolute_range("B", benchmark_lookup_header_row + 1, benchmark_lookup_header_row + max(len(benchmark_rows), 1)),
    )


def _build_calc_sheet(wb: Workbook, sheet_name: str, description: str) -> None:
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = sheet_name
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = description
    ws["A4"] = "Reserved for later formulas and controlled helper ranges."
    ws.column_dimensions["A"].width = 72


def _build_source_sheet(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    named_ranges: dict[str, str] | None = None,
) -> None:
    ws = wb.create_sheet(sheet_name)
    _write_dataframe(ws, df)
    ws.sheet_state = "hidden"

    header_to_index = {str(value).strip(): index for index, value in enumerate(df.columns, start=1)}
    data_end_row = max(len(df) + 1, 2)

    if named_ranges:
        for range_name, column_name in named_ranges.items():
            column_index = header_to_index[column_name]
            column_letter = get_column_letter(column_index)
            _add_named_range(wb, range_name, ws.title, _absolute_range(column_letter, 2, data_end_row))


def _build_overview_source_sheets(wb: Workbook, sheets: dict[str, pd.DataFrame]) -> None:
    kpi_summary = sheets["KPI_Summary"].copy()
    period_returns = sheets["Period_Returns"].copy()
    build_info = sheets["Build_Info"].copy()

    _require_columns(kpi_summary, "KPI_Summary", OVERVIEW_KPI_COLUMNS)
    _require_columns(period_returns, "Period_Returns", OVERVIEW_PERIOD_RETURN_COLUMNS)
    _require_columns(build_info, "Build_Info", ["Date_Max"])

    kpi_summary["Lookup_Key"] = (
        kpi_summary["Series_ID"].fillna("").astype(str).str.strip()
        + "|"
        + kpi_summary["Period"].fillna("").astype(str).str.strip()
    )

    _build_source_sheet(
        wb,
        "Source_KPI_Summary",
        kpi_summary,
        named_ranges={
            "Source_KPI_Lookup_Key": "Lookup_Key",
            "Source_KPI_Display_Name": "Display_Name",
            "Source_KPI_Series_Type": "Series_Type",
            "Source_KPI_Return_Total": "Return_Total",
            "Source_KPI_CAGR": "CAGR",
            "Source_KPI_Vol": "Vol",
            "Source_KPI_Sharpe": "Sharpe",
            "Source_KPI_Sortino": "Sortino",
            "Source_KPI_Max_DD": "Max_DD",
            "Source_KPI_Calmar": "Calmar",
        },
    )
    _build_source_sheet(
        wb,
        "Source_Period_Returns",
        period_returns,
        named_ranges={
            "Source_Period_Series_ID": "Series_ID",
            "Source_Period_30D": "30D",
            "Source_Period_YTD": "YTD",
            "Source_Period_1Y": "1Y",
            "Source_Period_Since_Start": "Since_Start",
        },
    )
    _build_source_sheet(
        wb,
        "Source_Build_Info",
        build_info,
        named_ranges={"Source_Build_Date_Max": "Date_Max"},
    )


def _build_performance_source_sheets(wb: Workbook, sheets: dict[str, pd.DataFrame]) -> None:
    chart_idx_wide = sheets["Chart_IDX_Wide"].copy()
    chart_dd_wide = sheets["Chart_DD_Wide"].copy()

    _chart_series_columns(chart_idx_wide, "Chart_IDX_Wide")
    _chart_series_columns(chart_dd_wide, "Chart_DD_Wide")

    _build_source_sheet(wb, "Source_Chart_IDX_Wide", chart_idx_wide)
    _build_source_sheet(wb, "Source_Chart_DD_Wide", chart_dd_wide)


def _build_structure_source_sheets(wb: Workbook, sheets: dict[str, pd.DataFrame]) -> None:
    allocation_snapshot = sheets["Allocation_Snapshot"].copy()
    _require_columns(
        allocation_snapshot,
        "Allocation_Snapshot",
        [column for column in ALLOCATION_SNAPSHOT_COLUMNS if column != "Display_Name"],
    )
    allocation_snapshot["Portfolio_Name"] = allocation_snapshot["Portfolio_Name"].fillna("").astype(str).str.strip()
    allocation_snapshot["Variant"] = allocation_snapshot["Variant"].fillna("").astype(str).str.strip().str.upper()
    if "Display_Name" not in allocation_snapshot.columns:
        allocation_snapshot["Display_Name"] = allocation_snapshot["Yahoo_Ticker"]
    allocation_snapshot["Display_Name"] = allocation_snapshot["Display_Name"].fillna("").astype(str).str.strip()
    allocation_snapshot["Yahoo_Ticker"] = allocation_snapshot["Yahoo_Ticker"].fillna("").astype(str).str.strip()
    allocation_snapshot["Display_Name"] = allocation_snapshot["Display_Name"].where(
        allocation_snapshot["Display_Name"] != "",
        allocation_snapshot["Yahoo_Ticker"],
    )
    allocation_snapshot["Weight"] = pd.to_numeric(allocation_snapshot["Weight"], errors="coerce")
    allocation_snapshot = allocation_snapshot.sort_values(
        ["Portfolio_Name", "Variant", "Weight", "Display_Name", "Yahoo_Ticker"],
        ascending=[True, True, False, True, True],
        na_position="last",
    ).reset_index(drop=True)
    allocation_snapshot["Source_Row_Index"] = range(1, len(allocation_snapshot) + 1)
    allocation_snapshot["Selection_Rank"] = (
        allocation_snapshot.groupby(["Portfolio_Name", "Variant"]).cumcount() + 1
    )
    allocation_snapshot["Lookup_Key"] = (
        allocation_snapshot["Portfolio_Name"]
        + "|"
        + allocation_snapshot["Variant"]
        + "|"
        + allocation_snapshot["Selection_Rank"].astype(str)
    )
    _build_source_sheet(
        wb,
        "Source_Allocation_Snapshot",
        allocation_snapshot,
        named_ranges={
            "Source_Allocation_Portfolio_Name": "Portfolio_Name",
            "Source_Allocation_Series_ID": "Series_ID",
            "Source_Allocation_Variant": "Variant",
            "Source_Allocation_Display_Name": "Display_Name",
            "Source_Allocation_Yahoo_Ticker": "Yahoo_Ticker",
            "Source_Allocation_Weight": "Weight",
            "Source_Allocation_Weight_Source": "Weight_Source",
            "Source_Allocation_Row_Index": "Source_Row_Index",
            "Source_Allocation_Lookup_Key": "Lookup_Key",
        },
    )


def _build_visible_sheet(wb: Workbook, sheet_name: str, description: str, defaults: dict[str, str]) -> None:
    ws = wb.create_sheet(sheet_name)
    _set_title(ws, sheet_name, description)
    status_rows = [
        ["Selected portfolio", "=Control_Selected_Portfolio"],
        ["Selected variant", "=Control_Selected_Variant"],
        ["Selected period", "=Control_Selected_Period"],
        ["Compare 1", "=Control_Selected_Compare_1"],
        ["Compare 2", "=Control_Selected_Compare_2"],
    ]
    if sheet_name == "Category":
        status_rows = [
            ["Selected category portfolio", defaults["selected_category_portfolio"]],
            ["Selected category variant", "REAL"],
            ["Selected category period", defaults["selected_category_period"]],
            ["Category 1", ""],
            ["Category 2", ""],
            ["Category 3", ""],
        ]
    _write_table(ws, 7, ["Field", "Value"], status_rows)
    ws["A16"] = "Next step"
    ws["A16"].font = Font(bold=True)
    ws["A17"] = "This sheet currently exposes only workbook structure and control placeholders."
    ws["A18"] = "Later threads should add controlled formulas, tables and charts on top of this contract."
    ws.column_dimensions["B"].width = 24


def _overview_kpi_lookup_formula(series_id_ref: str, metric_range_name: str) -> str:
    return (
        f'=IF({series_id_ref}="","",'
        f'IFERROR(INDEX({metric_range_name},MATCH({series_id_ref}&"|"&Calc_Main_Selected_Period,'
        f'Source_KPI_Lookup_Key,0)),""))'
    )


def _overview_period_lookup_formula(series_id_ref: str, metric_range_name: str) -> str:
    return (
        f'=IF({series_id_ref}="","",'
        f'IFERROR(INDEX({metric_range_name},MATCH({series_id_ref},Source_Period_Series_ID,0)),""))'
    )


def _category_kpi_lookup_formula(series_id_ref: str, metric_range_name: str) -> str:
    return (
        f'=IF({series_id_ref}="","",'
        f'IFERROR(INDEX({metric_range_name},MATCH({series_id_ref}&"|"&Calc_Category_Selected_Period,'
        f'Source_KPI_Lookup_Key,0)),""))'
    )


def _category_period_lookup_formula(series_id_ref: str, metric_range_name: str) -> str:
    return (
        f'=IF({series_id_ref}="","",'
        f'IFERROR(INDEX({metric_range_name},MATCH({series_id_ref},Source_Period_Series_ID,0)),""))'
    )


def _performance_chart_formula(series_id_ref: str, source_sheet_name: str, source_col_ref: str, data_row_ref: str) -> str:
    return (
        f'=IF({series_id_ref}="","",'
        f'IFERROR(INDEX({_quote_sheet_name(source_sheet_name)}!$A:$ZZ,'
        f'{data_row_ref}+1,MATCH({series_id_ref},{source_col_ref},0)),""))'
    )


def _selected_label_formula(label_ref: str, series_id_ref: str) -> str:
    return f'=IF({series_id_ref}="","",{label_ref})'


def _build_performance_chart(
    ws,
    title: str,
    top_left_cell: str,
    category_col: int,
    min_col: int,
    max_col: int,
    min_row: int,
    max_row: int,
) -> None:
    chart = LineChart()
    chart.title = title
    chart.style = 2
    chart.height = 7
    chart.width = 18
    chart.y_axis.title = ""
    chart.legend.position = "r"
    data = Reference(ws, min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row)
    categories = Reference(ws, min_col=category_col, max_col=category_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, top_left_cell)


def _build_structure_chart(ws, top_left_cell: str, min_row: int, max_row: int) -> None:
    chart = BarChart()
    chart.type = "bar"
    chart.style = 2
    chart.title = "Weight distribution"
    chart.height = 7
    chart.width = 14
    chart.y_axis.title = ""
    chart.x_axis.title = ""
    chart.legend = None
    data = Reference(ws, min_col=10, max_col=10, min_row=min_row, max_row=max_row)
    categories = Reference(ws, min_col=9, max_col=9, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, top_left_cell)


def _build_overview_sheet(wb: Workbook, defaults: dict[str, str]) -> None:
    ws = wb.create_sheet("Overview")
    ws["A1"] = "Overview"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Shared v1 dashboard view with the main editable controls."

    control_headers = [
        "Primary portfolio",
        "Variant",
        "Period",
        "Compare 1",
        "Compare 2",
        "",
        "Data through",
    ]
    control_values = [
        defaults["selected_portfolio"],
        defaults["selected_variant"],
        defaults["selected_period"],
        "",
        "",
        "",
        '=IFERROR(INDEX(Source_Build_Date_Max,1),"")',
    ]
    for column_index, value in enumerate(control_headers, start=1):
        cell = ws.cell(row=4, column=column_index, value=value)
        cell.font = Font(bold=True)
    for column_index, value in enumerate(control_values, start=1):
        ws.cell(row=5, column=column_index, value=value)
    ws["G5"].number_format = "yyyy-mm-dd"
    ws["A6"] = "Edit selections here. Performance and Structure mirror the same main controls."

    for range_name, cell_ref in MAIN_CONTROL_SOURCE_CELLS.items():
        _add_named_range(wb, range_name, ws.title, cell_ref)
    _add_list_validation(ws, "A5", "=List_Main_Portfolio")
    _add_list_validation(ws, "B5", "=List_Variant")
    _add_list_validation(ws, "C5", "=List_Period")
    for cell_ref in ["D5", "E5"]:
        _add_list_validation(ws, cell_ref, "=List_Compare_Series")

    ws["A7"] = "Primary KPI"
    ws["A7"].font = Font(bold=True)
    ws["B7"] = "=Calc_Main_Primary_Display_Name"
    ws["A8"] = "Metric"
    ws["B8"] = "Value"
    ws["A8"].font = Font(bold=True)
    ws["B8"].font = Font(bold=True)

    primary_series_ref = "Calc_Main_Primary_Series_ID"
    kpi_rows = [
        ("Return", _overview_kpi_lookup_formula(primary_series_ref, "Source_KPI_Return_Total")),
        ("CAGR", _overview_kpi_lookup_formula(primary_series_ref, "Source_KPI_CAGR")),
        ("Vol", _overview_kpi_lookup_formula(primary_series_ref, "Source_KPI_Vol")),
        ("Sharpe", _overview_kpi_lookup_formula(primary_series_ref, "Source_KPI_Sharpe")),
        ("Max DD", _overview_kpi_lookup_formula(primary_series_ref, "Source_KPI_Max_DD")),
        ("Calmar", _overview_kpi_lookup_formula(primary_series_ref, "Source_KPI_Calmar")),
    ]
    for row_index, (label, formula) in enumerate(kpi_rows, start=9):
        ws.cell(row=row_index, column=1, value=label)
        ws.cell(row=row_index, column=2, value=formula)

    ws["D7"] = "Comparison"
    ws["D7"].font = Font(bold=True)
    comparison_headers = [
        "Role",
        "Series",
        "Series_ID",
        "Type",
        "Period",
        "Return",
        "CAGR",
        "Vol",
        "Sharpe",
        "Sortino",
        "Max_DD",
        "Calmar",
    ]
    for column_index, header in enumerate(comparison_headers, start=4):
        cell = ws.cell(row=8, column=column_index, value=header)
        cell.font = Font(bold=True)

    for row_offset, calc_row in enumerate(OVERVIEW_ROLE_ROWS.values(), start=9):
        series_id_ref = f"$F{row_offset}"
        ws[f"D{row_offset}"] = _selected_label_formula(f"'Calc_Main'!D{calc_row}", series_id_ref)
        ws[f"E{row_offset}"] = _selected_label_formula(f"'Calc_Main'!E{calc_row}", series_id_ref)
        ws[f"F{row_offset}"] = f"='Calc_Main'!F{calc_row}"
        ws[f"G{row_offset}"] = _overview_kpi_lookup_formula(series_id_ref, "Source_KPI_Series_Type")
        ws[f"H{row_offset}"] = f'=IF({series_id_ref}="","",Calc_Main_Selected_Period)'
        ws[f"I{row_offset}"] = _overview_kpi_lookup_formula(series_id_ref, "Source_KPI_Return_Total")
        ws[f"J{row_offset}"] = _overview_kpi_lookup_formula(series_id_ref, "Source_KPI_CAGR")
        ws[f"K{row_offset}"] = _overview_kpi_lookup_formula(series_id_ref, "Source_KPI_Vol")
        ws[f"L{row_offset}"] = _overview_kpi_lookup_formula(series_id_ref, "Source_KPI_Sharpe")
        ws[f"M{row_offset}"] = _overview_kpi_lookup_formula(series_id_ref, "Source_KPI_Sortino")
        ws[f"N{row_offset}"] = _overview_kpi_lookup_formula(series_id_ref, "Source_KPI_Max_DD")
        ws[f"O{row_offset}"] = _overview_kpi_lookup_formula(series_id_ref, "Source_KPI_Calmar")

    ws["A17"] = "Period Returns"
    ws["A17"].font = Font(bold=True)
    period_headers = ["Role", "Series", "Series_ID", "30D", "YTD", "1Y", "Since_Start"]
    for column_index, header in enumerate(period_headers, start=1):
        cell = ws.cell(row=18, column=column_index, value=header)
        cell.font = Font(bold=True)

    for row_offset, calc_row in enumerate(OVERVIEW_ROLE_ROWS.values(), start=19):
        series_id_ref = f"$C{row_offset}"
        ws[f"A{row_offset}"] = _selected_label_formula(f"'Calc_Main'!D{calc_row}", series_id_ref)
        ws[f"B{row_offset}"] = _selected_label_formula(f"'Calc_Main'!E{calc_row}", series_id_ref)
        ws[f"C{row_offset}"] = f"='Calc_Main'!F{calc_row}"
        ws[f"D{row_offset}"] = _overview_period_lookup_formula(series_id_ref, "Source_Period_30D")
        ws[f"E{row_offset}"] = _overview_period_lookup_formula(series_id_ref, "Source_Period_YTD")
        ws[f"F{row_offset}"] = _overview_period_lookup_formula(series_id_ref, "Source_Period_1Y")
        ws[f"G{row_offset}"] = _overview_period_lookup_formula(series_id_ref, "Source_Period_Since_Start")

    for cell_ref in ["B9", "B10", "B11", "B13", "I9", "J9", "K9", "N9", "D19", "E19", "F19", "G19"]:
        ws[cell_ref].number_format = "0.00%"
    for cell_ref in ["B12", "B14", "L9", "M9", "O9"]:
        ws[cell_ref].number_format = "0.00"
    for row_index in range(9, 13):
        for column_letter in ["I", "J", "K", "N"]:
            ws[f"{column_letter}{row_index}"].number_format = "0.00%"
        for column_letter in ["L", "M", "O"]:
            ws[f"{column_letter}{row_index}"].number_format = "0.00"
    for row_index in range(19, 23):
        for column_letter in ["D", "E", "F", "G"]:
            ws[f"{column_letter}{row_index}"].number_format = "0.00%"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 22
    for column_letter in ["G", "H", "I", "J", "K", "L", "M", "N", "O"]:
        ws.column_dimensions[column_letter].width = 14
    ws.column_dimensions["F"].hidden = True
    ws.column_dimensions["C"].hidden = True
    ws.freeze_panes = "A4"


def _build_structure_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Structure")
    ws["A1"] = "Structure"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Current allocation snapshot driven by the shared main controls."

    status_headers = [
        "Primary portfolio",
        "Variant",
        "Selection status",
        "Snapshot as-of",
        "Edit controls",
    ]
    status_formulas = [
        "=Control_Selected_Portfolio",
        "=Calc_Main_Structure_Variant",
        "=Calc_Main_Structure_Status",
        '=IFERROR(INDEX(Source_Build_Date_Max,1),"")',
    ]
    for column_index, value in enumerate(status_headers, start=1):
        ws.cell(row=4, column=column_index, value=value).font = Font(bold=True)
    for column_index, value in enumerate(status_formulas, start=1):
        ws.cell(row=5, column=column_index, value=value)
    ws["E5"].number_format = "yyyy-mm-dd"
    _add_sheet_link(ws, "E5", "Overview", "A4", "Open main controls")
    ws["A6"] = "Main controls are edited on Overview and reused here."

    ws["A7"] = "Snapshot summary"
    ws["A7"].font = Font(bold=True)
    summary_rows = [
        ("Selected series", "=Calc_Main_Primary_Display_Name"),
        ("Allocation rows", "=Calc_Main_Structure_Row_Count"),
        ("Weight sum", "=Calc_Main_Structure_Weight_Sum"),
        ("Deviation vs 100%", "=Calc_Main_Structure_Weight_Deviation"),
    ]
    for row_index, (label, formula) in enumerate(summary_rows, start=8):
        ws.cell(row=row_index, column=1, value=label)
        ws.cell(row=row_index, column=2, value=formula)

    ws["D7"] = "Allocation table"
    ws["D7"].font = Font(bold=True)
    table_headers = ["Ticker", "Weight", "Weight_Source", "Series_ID"]
    for column_index, header in enumerate(table_headers, start=4):
        ws.cell(row=8, column=column_index, value=header).font = Font(bold=True)

    for row_index in range(9, 9 + STRUCTURE_TABLE_MAX_ROWS):
        calc_row = CALC_MAIN_STRUCTURE_TABLE_FIRST_ROW + (row_index - 9)
        ws[f"D{row_index}"] = f"='Calc_Main'!X{calc_row}"
        ws[f"E{row_index}"] = f"='Calc_Main'!Y{calc_row}"
        ws[f"F{row_index}"] = f"='Calc_Main'!Z{calc_row}"
        ws[f"G{row_index}"] = f"='Calc_Main'!AA{calc_row}"

    ws["I7"] = "Chart helper"
    ws["I7"].font = Font(bold=True)
    ws["I8"] = "Ticker"
    ws["J8"] = "Weight"
    ws["I8"].font = Font(bold=True)
    ws["J8"].font = Font(bold=True)
    for row_index in range(9, 9 + STRUCTURE_TOP_COUNT):
        calc_row = CALC_MAIN_STRUCTURE_TOP_FIRST_ROW + (row_index - 9)
        ws[f"I{row_index}"] = f"='Calc_Main'!X{calc_row}"
        ws[f"J{row_index}"] = f"='Calc_Main'!Y{calc_row}"

    _build_structure_chart(ws, "I12", 8, 8 + STRUCTURE_TOP_COUNT)

    ws["B9"].number_format = "0"
    ws["B10"].number_format = "0.00%"
    ws["B11"].number_format = "0.00%"
    for row_index in range(9, 9 + STRUCTURE_TABLE_MAX_ROWS):
        ws[f"E{row_index}"].number_format = "0.00%"
    for row_index in range(9, 9 + STRUCTURE_TOP_COUNT):
        ws[f"J{row_index}"].number_format = "0.00%"

    for column_letter, width in {
        "A": 18,
        "B": 24,
        "C": 16,
        "D": 12,
        "E": 18,
        "G": 22,
        "I": 16,
        "J": 12,
    }.items():
        ws.column_dimensions[column_letter].width = width
    ws.column_dimensions["G"].hidden = True
    ws.freeze_panes = "A4"


def _build_performance_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Performance")
    ws["A1"] = "Performance"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Shared v1 performance view driven by the shared main controls."

    status_headers = [
        "Primary portfolio",
        "Variant",
        "Period",
        "Compare 1",
        "Compare 2",
        "Data through",
        "Edit controls",
    ]
    status_formulas = [
        "=Control_Selected_Portfolio",
        "=Control_Selected_Variant",
        "=Control_Selected_Period",
        "=Control_Selected_Compare_1",
        "=Control_Selected_Compare_2",
        '=IFERROR(INDEX(Source_Build_Date_Max,1),"")',
    ]
    for column_index, value in enumerate(status_headers, start=1):
        ws.cell(row=4, column=column_index, value=value).font = Font(bold=True)
    for column_index, value in enumerate(status_formulas, start=1):
        ws.cell(row=5, column=column_index, value=value)
    ws["F5"].number_format = "yyyy-mm-dd"
    _add_sheet_link(ws, "G5", "Overview", "A4", "Open main controls")
    ws["A6"] = "Main controls are edited on Overview and reused here."

    ws["A7"] = "Active selection"
    ws["A7"].font = Font(bold=True)
    selection_headers = ["Role", "Series", "Series_ID", "Period"]
    for column_index, header in enumerate(selection_headers, start=1):
        ws.cell(row=8, column=column_index, value=header).font = Font(bold=True)
    for row_offset, calc_row in enumerate(PERFORMANCE_ROLE_ROWS.values(), start=9):
        series_id_ref = f"$C{row_offset}"
        ws[f"A{row_offset}"] = _selected_label_formula(f"'Calc_Main'!H{calc_row}", series_id_ref)
        ws[f"B{row_offset}"] = _selected_label_formula(f"'Calc_Main'!I{calc_row}", series_id_ref)
        ws[f"C{row_offset}"] = f"='Calc_Main'!J{calc_row}"
        ws[f"D{row_offset}"] = f'=IF({series_id_ref}="","",Calc_Main_Selected_Period)'

    ws["F7"] = "Summary"
    ws["F7"].font = Font(bold=True)
    summary_headers = ["Role", "Series", "30D", "YTD", "1Y", "Since_Start", "Vol", "Max_DD", "Sharpe"]
    for column_index, header in enumerate(summary_headers, start=6):
        ws.cell(row=8, column=column_index, value=header).font = Font(bold=True)
    for row_offset, calc_row in enumerate(PERFORMANCE_ROLE_ROWS.values(), start=9):
        series_id_ref = f"$H{row_offset}"
        ws[f"F{row_offset}"] = _selected_label_formula(f"'Calc_Main'!H{calc_row}", series_id_ref)
        ws[f"G{row_offset}"] = _selected_label_formula(f"'Calc_Main'!I{calc_row}", series_id_ref)
        ws[f"H{row_offset}"] = f"='Calc_Main'!J{calc_row}"
        ws[f"I{row_offset}"] = _overview_period_lookup_formula(series_id_ref, "Source_Period_30D")
        ws[f"J{row_offset}"] = _overview_period_lookup_formula(series_id_ref, "Source_Period_YTD")
        ws[f"K{row_offset}"] = _overview_period_lookup_formula(series_id_ref, "Source_Period_1Y")
        ws[f"L{row_offset}"] = _overview_period_lookup_formula(series_id_ref, "Source_Period_Since_Start")
        ws[f"M{row_offset}"] = _overview_kpi_lookup_formula(series_id_ref, "Source_KPI_Vol")
        ws[f"N{row_offset}"] = _overview_kpi_lookup_formula(series_id_ref, "Source_KPI_Max_DD")
        ws[f"O{row_offset}"] = _overview_kpi_lookup_formula(series_id_ref, "Source_KPI_Sharpe")

    ws["A15"] = "Index development"
    ws["A15"].font = Font(bold=True)
    performance_chart_headers = {
        "A16": "Date",
        "B16": _selected_label_formula("'Calc_Main'!I18", "'Calc_Main'!J18"),
        "C16": _selected_label_formula("'Calc_Main'!I19", "'Calc_Main'!J19"),
        "D16": _selected_label_formula("'Calc_Main'!I20", "'Calc_Main'!J20"),
        "G16": "Date",
        "H16": _selected_label_formula("'Calc_Main'!I18", "'Calc_Main'!J18"),
        "I16": _selected_label_formula("'Calc_Main'!I19", "'Calc_Main'!J19"),
        "J16": _selected_label_formula("'Calc_Main'!I20", "'Calc_Main'!J20"),
    }
    for cell_ref, value in performance_chart_headers.items():
        ws[cell_ref] = value
        ws[cell_ref].font = Font(bold=True)

    calc_main_max_row = wb["Calc_Main"].max_row
    performance_chart_max_row = calc_main_max_row - 1
    dd_data_start_col = 7
    for row_index in range(17, performance_chart_max_row + 1):
        calc_row = row_index + 1
        ws[f"A{row_index}"] = f"='Calc_Main'!L{calc_row}"
        ws[f"B{row_index}"] = f"='Calc_Main'!M{calc_row}"
        ws[f"C{row_index}"] = f"='Calc_Main'!N{calc_row}"
        ws[f"D{row_index}"] = f"='Calc_Main'!O{calc_row}"
        ws.cell(row=row_index, column=dd_data_start_col, value=f"='Calc_Main'!Q{calc_row}")
        ws.cell(row=row_index, column=dd_data_start_col + 1, value=f"='Calc_Main'!R{calc_row}")
        ws.cell(row=row_index, column=dd_data_start_col + 2, value=f"='Calc_Main'!S{calc_row}")
        ws.cell(row=row_index, column=dd_data_start_col + 3, value=f"='Calc_Main'!T{calc_row}")

    ws["G15"] = "Drawdown"
    ws["G15"].font = Font(bold=True)

    _build_performance_chart(ws, "Index development", "A23", 1, 2, 4, 16, performance_chart_max_row)
    _build_performance_chart(ws, "Drawdown", "J23", 7, 8, 10, 16, performance_chart_max_row)

    for row_index in range(9, 12):
        for column_letter in ["I", "J", "K", "L", "M", "N"]:
            ws[f"{column_letter}{row_index}"].number_format = "0.00%"
        ws[f"O{row_index}"].number_format = "0.00"
    for row_index in range(17, performance_chart_max_row + 1):
        for column_letter in ["B", "C", "D", "H", "I", "J"]:
            ws[f"{column_letter}{row_index}"].number_format = "0.00%"
        ws[f"A{row_index}"].number_format = "yyyy-mm-dd"
        ws[f"G{row_index}"].number_format = "yyyy-mm-dd"

    for column_letter, width in {
        "A": 14,
        "B": 24,
        "C": 20,
        "D": 14,
        "F": 14,
        "G": 24,
        "H": 20,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 14,
        "M": 12,
        "N": 12,
        "O": 12,
    }.items():
        ws.column_dimensions[column_letter].width = width
    ws.column_dimensions["C"].hidden = True
    ws.column_dimensions["H"].hidden = True
    ws.freeze_panes = "A4"


def _build_calc_main_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Calc_Main")
    ws["A1"] = "Calc_Main"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Controlled helper area for the shared main dashboard state."
    ws["A4"] = "Key"
    ws["B4"] = "Value / Formula"
    ws["A4"].font = Font(bold=True)
    ws["B4"].font = Font(bold=True)

    helper_rows = [
        [
            "Primary_Display_Name",
            '=IFERROR(INDEX(Lookup_Main_Series_Display_Name,'
            'MATCH(Control_Selected_Portfolio&"|"&Control_Selected_Variant,Lookup_Main_Series_Key,0)),"")',
        ],
        [
            "Primary_Series_ID",
            '=IFERROR(INDEX(Lookup_Main_Series_ID,'
            'MATCH(Control_Selected_Portfolio&"|"&Control_Selected_Variant,Lookup_Main_Series_Key,0)),"")',
        ],
        ["Selected_Period", "=Control_Selected_Period"],
        ["Compare_1_Display_Name", "=Control_Selected_Compare_1"],
        [
            "Compare_1_Series_ID",
            '=IF(Control_Selected_Compare_1="","",IFERROR(INDEX(Lookup_Compare_Series_ID,'
            'MATCH(Control_Selected_Compare_1,Lookup_Compare_Display_Name,0)),""))',
        ],
        ["Compare_2_Display_Name", "=Control_Selected_Compare_2"],
        [
            "Compare_2_Series_ID",
            '=IF(Control_Selected_Compare_2="","",IFERROR(INDEX(Lookup_Compare_Series_ID,'
            'MATCH(Control_Selected_Compare_2,Lookup_Compare_Display_Name,0)),""))',
        ],
    ]
    _write_table(ws, 5, ["Helper_Key", "Value"], helper_rows)

    ws["D4"] = "Overview_Base"
    ws["D4"].font = Font(bold=True)
    ws["D5"] = "Role"
    ws["E5"] = "Display_Name"
    ws["F5"] = "Series_ID"
    for cell_ref in ["D5", "E5", "F5"]:
        ws[cell_ref].font = Font(bold=True)

    overview_rows = [
        ["Primary", "=B6", "=B7"],
        ["Compare_1", "=B9", "=B10"],
        ["Compare_2", "=B11", "=B12"],
    ]
    for row_index, values in enumerate(overview_rows, start=6):
        ws.cell(row=row_index, column=4, value=values[0])
        ws.cell(row=row_index, column=5, value=values[1])
        ws.cell(row=row_index, column=6, value=values[2])

    ws["H4"] = "Performance_Selected"
    ws["H4"].font = Font(bold=True)
    performance_headers = ["Role", "Display_Name", "Series_ID"]
    for column_index, header in enumerate(performance_headers, start=8):
        ws.cell(row=17, column=column_index, value=header).font = Font(bold=True)
    performance_rows = [
        ["Primary", "=B6", "=B7"],
        ["Compare_1", "=B9", "=B10"],
        ["Compare_2", "=B11", "=B12"],
    ]
    for row_index, values in enumerate(performance_rows, start=18):
        for column_index, value in enumerate(values, start=8):
            ws.cell(row=row_index, column=column_index, value=value)

    ws["L4"] = "Performance_IDX_Helper"
    ws["L4"].font = Font(bold=True)
    idx_headers = ["Date", "Primary", "Compare_1", "Compare_2"]
    for column_index, header in enumerate(idx_headers, start=12):
        ws.cell(row=17, column=column_index, value=header).font = Font(bold=True)

    ws["Q4"] = "Performance_DD_Helper"
    ws["Q4"].font = Font(bold=True)
    for column_index, header in enumerate(idx_headers, start=17):
        ws.cell(row=17, column=column_index, value=header).font = Font(bold=True)

    idx_row_count = max(wb["Source_Chart_IDX_Wide"].max_row - 1, 1)
    dd_row_count = max(wb["Source_Chart_DD_Wide"].max_row - 1, 1)
    row_count = max(idx_row_count, dd_row_count)

    for row_index in range(18, 18 + row_count):
        source_data_row_formula = f"ROW()-17"
        ws.cell(
            row=row_index,
            column=12,
            value=f'=IFERROR(INDEX(Source_Chart_IDX_Wide!$A:$A,{source_data_row_formula}+1),"")',
        )
        ws.cell(
            row=row_index,
            column=17,
            value=f'=IFERROR(INDEX(Source_Chart_DD_Wide!$A:$A,{source_data_row_formula}+1),"")',
        )

        idx_series_refs = ["$J$18", "$J$19", "$J$20"]
        dd_series_refs = idx_series_refs
        for offset, series_id_ref in enumerate(idx_series_refs, start=13):
            ws.cell(
                row=row_index,
                column=offset,
                value=_performance_chart_formula(
                    series_id_ref,
                    "Source_Chart_IDX_Wide",
                    "Source_Chart_IDX_Wide!$1:$1",
                    source_data_row_formula,
                ),
            )
        for offset, series_id_ref in enumerate(dd_series_refs, start=18):
            ws.cell(
                row=row_index,
                column=offset,
                value=_performance_chart_formula(
                    series_id_ref,
                    "Source_Chart_DD_Wide",
                    "Source_Chart_DD_Wide!$1:$1",
                    source_data_row_formula,
                ),
            )

    allocation_source_last_row = max(wb["Source_Allocation_Snapshot"].max_row, 2)
    allocation_portfolio_ref = f"Source_Allocation_Snapshot!$A$2:$A${allocation_source_last_row}"
    allocation_series_id_ref = f"Source_Allocation_Snapshot!$B$2:$B${allocation_source_last_row}"
    allocation_variant_ref = f"Source_Allocation_Snapshot!$C$2:$C${allocation_source_last_row}"
    allocation_display_name_ref = f"Source_Allocation_Snapshot!$D$2:$D${allocation_source_last_row}"
    allocation_ticker_ref = f"Source_Allocation_Snapshot!$E$2:$E${allocation_source_last_row}"
    allocation_weight_ref = f"Source_Allocation_Snapshot!$F$2:$F${allocation_source_last_row}"
    allocation_weight_source_ref = f"Source_Allocation_Snapshot!$G$2:$G${allocation_source_last_row}"
    allocation_row_index_ref = f"Source_Allocation_Snapshot!$H$2:$H${allocation_source_last_row}"

    ws["W4"] = "Structure_Helper"
    ws["W4"].font = Font(bold=True)
    structure_helper_rows = [
        ["Selected_Portfolio", "=Control_Selected_Portfolio"],
        [
            "Selected_Variant",
            f'=IF(COUNTIFS({allocation_portfolio_ref},$X$6,{allocation_variant_ref},Control_Selected_Variant)>0,'
            'Control_Selected_Variant,'
            f'IF(COUNTIFS({allocation_portfolio_ref},$X$6,{allocation_variant_ref},"CUR")>0,"CUR",'
            f'IF(COUNTIFS({allocation_portfolio_ref},$X$6,{allocation_variant_ref},"REAL")>0,"REAL",'
            f'IF(COUNTIFS({allocation_portfolio_ref},$X$6,{allocation_variant_ref},"TGT")>0,"TGT",""))))',
        ],
        [
            "Status",
            '=IF($X$6="","",IF($X$7="","No allocation rows for portfolio",'
            'IF($X$7=Control_Selected_Variant,"Allocation snapshot available",'
            'IF(Control_Selected_Variant="REAL",'
            '"REAL allocation snapshot missing in source data; showing "&$X$7&" fallback",'
            '"Allocation snapshot available via "&$X$7&" fallback"))))',
        ],
        ["Row_Count", f'=COUNTIFS({allocation_portfolio_ref},$X$6,{allocation_variant_ref},$X$7)'],
        ["Weight_Sum", f'=SUMIFS({allocation_weight_ref},{allocation_portfolio_ref},$X$6,{allocation_variant_ref},$X$7)'],
        ["Weight_Deviation", '=IF($X$9=0,"",ABS(1-$X$10))'],
    ]
    for row_index, values in enumerate(structure_helper_rows, start=6):
        ws.cell(row=row_index, column=23, value=values[0])
        ws.cell(row=row_index, column=24, value=values[1])

    ws["X14"] = "Snapshot as-of"
    ws["Y14"] = '=IFERROR(INDEX(Source_Build_Date_Max,1),"")'
    ws["W16"] = "Structure_Filtered_Allocation"
    ws["W16"].font = Font(bold=True)
    for column_index, header in enumerate(["Match_Row", "Display_Name", "Weight", "Weight_Source", "Series_ID"], start=23):
        ws.cell(row=17, column=column_index, value=header).font = Font(bold=True)

    allocation_first_row = 18
    allocation_last_row = allocation_first_row + STRUCTURE_TABLE_MAX_ROWS - 1
    for row_index in range(allocation_first_row, allocation_last_row + 1):
        rank_formula = f"ROWS($W${allocation_first_row}:W{row_index})"
        match_formula = (
            '=IFERROR(MATCH('
            f'$X$6&"|"&$X$7&"|"&{rank_formula},'
            'Source_Allocation_Lookup_Key,0),"")'
        )
        row_ref = f"$W{row_index}"
        ws[f"W{row_index}"] = match_formula
        ws[f"X{row_index}"] = f'=IF({row_ref}="","",INDEX({allocation_display_name_ref},{row_ref}))'
        ws[f"Y{row_index}"] = f'=IF({row_ref}="","",INDEX({allocation_weight_ref},{row_ref}))'
        ws[f"Z{row_index}"] = f'=IF({row_ref}="","",INDEX({allocation_weight_source_ref},{row_ref}))'
        ws[f"AA{row_index}"] = f'=IF({row_ref}="","",INDEX({allocation_series_id_ref},{row_ref}))'

    ws["W44"] = "Structure_Top_Holdings"
    ws["W44"].font = Font(bold=True)
    ws["X45"] = "Display_Name"
    ws["Y45"] = "Weight"
    ws["X45"].font = Font(bold=True)
    ws["Y45"].font = Font(bold=True)
    top_first_row = 46
    top_last_row = top_first_row + STRUCTURE_TOP_COUNT - 1
    for row_index in range(top_first_row, top_last_row + 1):
        source_row = allocation_first_row + (row_index - top_first_row)
        ws[f"W{row_index}"] = f'=IF($X{source_row}="","",ROW()-{top_first_row - 1})'
        ws[f"X{row_index}"] = f'=IF($X{source_row}="","",$X{source_row})'
        ws[f"Y{row_index}"] = f'=IF($X{source_row}="","",$Y{source_row})'

    for name, cell_ref in {
        "Calc_Main_Primary_Display_Name": "B6",
        "Calc_Main_Primary_Series_ID": "B7",
        "Calc_Main_Selected_Period": "B8",
        "Calc_Main_Overview_Base": "$D$6:$F$8",
        "Calc_Main_Performance_Selected": "$H$18:$J$20",
        "Calc_Main_Performance_IDX_Data": f"$L$17:$O${17 + row_count}",
        "Calc_Main_Performance_DD_Data": f"$Q$17:$T${17 + row_count}",
        "Calc_Main_Structure_Status": "$X$8",
        "Calc_Main_Structure_Variant": "$X$7",
        "Calc_Main_Structure_Row_Count": "$X$9",
        "Calc_Main_Structure_Weight_Sum": "$X$10",
        "Calc_Main_Structure_Weight_Deviation": "$X$11",
        "Calc_Main_Structure_As_Of": "$Y$14",
        "Calc_Main_Structure_Table": f"$X${allocation_first_row}:$AA${allocation_last_row}",
        "Calc_Main_Structure_Top_Data": f"$X$45:$Y${top_last_row}",
    }.items():
        _add_named_range(wb, name, ws.title, cell_ref)

    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 28
    ws.column_dimensions["J"].width = 22
    for column_letter in ["L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "W", "X", "Y", "Z", "AA"]:
        ws.column_dimensions[column_letter].width = 14
    ws.column_dimensions["W"].hidden = True
    ws.column_dimensions["AA"].hidden = True


def _build_calc_category_sheet(wb: Workbook, category_row_limit: int, benchmark_row_limit: int) -> None:
    ws = wb.create_sheet("Calc_Category")
    ws["A1"] = "Calc_Category"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Controlled helper area for category-specific REAL analysis."
    ws["A4"] = "Key"
    ws["B4"] = "Value / Formula"
    ws["A4"].font = Font(bold=True)
    ws["B4"].font = Font(bold=True)

    helper_rows = [
        ["Selected_Portfolio", "=Control_Selected_Category_Portfolio"],
        ["Selected_Variant", '=IF(Control_Selected_Category_Variant="","REAL","REAL")'],
        ["Selected_Period", "=Control_Selected_Category_Period"],
        ["Selected_Chart_Category_Display_Name", "=Control_Selected_Category_Chart"],
        [
            "Selected_Chart_Category_Series_ID",
            '=IF(Control_Selected_Category_Chart="","",IFERROR(INDEX(Lookup_Category_Series_ID,'
            'MATCH(Control_Selected_Category_Portfolio&"|"&Control_Selected_Category_Chart,Lookup_Category_Key,0)),""))',
        ],
        ["Selected_Benchmark_Display_Name", "=Control_Selected_Category_Benchmark"],
        [
            "Selected_Benchmark_Series_ID",
            '=IF(Control_Selected_Category_Benchmark="","",IFERROR(INDEX(Lookup_Benchmark_Series_ID,'
            'MATCH(Control_Selected_Category_Benchmark,Lookup_Benchmark_Display_Name,0)),""))',
        ],
        ["Available_Category_Count", "=COUNTIF($K$6:$K$1048576,\"<>\")"],
        ["Status", '=IF($B$6="","",IF($B$13=0,"No REAL category series available for portfolio","Category selection ready"))'],
    ]
    _write_table(ws, 5, ["Helper_Key", "Value"], helper_rows)

    ws["D4"] = "Chart_Selected"
    ws["D4"].font = Font(bold=True)
    for column_index, header in enumerate(["Role", "Display_Name", "Series_ID"], start=4):
        ws.cell(row=5, column=column_index, value=header).font = Font(bold=True)
    selected_rows = [
        ["Category", "=B9", "=B10"],
        ["Benchmark", "=B11", "=B12"],
    ]
    for row_index, values in enumerate(selected_rows, start=6):
        for column_index, value in enumerate(values, start=4):
            ws.cell(row=row_index, column=column_index, value=value)

    ws["J4"] = "Category_List_Helper"
    ws["J4"].font = Font(bold=True)
    for column_index, header in enumerate(["Match_Row", "Display_Name", "Series_ID"], start=10):
        ws.cell(row=5, column=column_index, value=header).font = Font(bold=True)
    list_first_row = 6
    list_last_row = list_first_row + category_row_limit - 1
    for row_index in range(list_first_row, list_last_row + 1):
        row_rank_formula = f'ROWS($K${list_first_row}:K{row_index})'
        match_formula = f'=IFERROR(MATCH($B$6&"|"&{row_rank_formula},Lookup_Category_Portfolio_Row_Key,0),"")'
        ws[f"J{row_index}"] = match_formula
        ws[f"K{row_index}"] = f'=IFERROR(INDEX(Lookup_Category_Display_Name,MATCH($B$6&"|"&{row_rank_formula},Lookup_Category_Portfolio_Row_Key,0)),"")'
        ws[f"L{row_index}"] = f'=IFERROR(INDEX(Lookup_Category_Series_ID,MATCH($B$6&"|"&{row_rank_formula},Lookup_Category_Portfolio_Row_Key,0)),"")'

    ws["N4"] = "Benchmark_List_Helper"
    ws["N4"].font = Font(bold=True)
    for column_index, header in enumerate(["Display_Name", "Series_ID"], start=14):
        ws.cell(row=5, column=column_index, value=header).font = Font(bold=True)
    benchmark_first_row = 6
    benchmark_last_row = benchmark_first_row + benchmark_row_limit - 1
    for row_index in range(benchmark_first_row, benchmark_last_row + 1):
        ws[f"N{row_index}"] = f'=IFERROR(INDEX(Lookup_Benchmark_Display_Name,ROWS($N${benchmark_first_row}:N{row_index})),"")'
        ws[f"O{row_index}"] = f'=IF($N{row_index}="","",IFERROR(INDEX(Lookup_Benchmark_Series_ID,ROWS($N${benchmark_first_row}:N{row_index})),""))'

    ws["Q4"] = "Category_IDX_Helper"
    ws["Q4"].font = Font(bold=True)
    for column_index, header in enumerate(["Date", "Category", "Benchmark"], start=17):
        ws.cell(row=5, column=column_index, value=header).font = Font(bold=True)

    ws["U4"] = "Category_DD_Helper"
    ws["U4"].font = Font(bold=True)
    for column_index, header in enumerate(["Date", "Category", "Benchmark"], start=21):
        ws.cell(row=5, column=column_index, value=header).font = Font(bold=True)

    idx_row_count = max(wb["Source_Chart_IDX_Wide"].max_row - 1, 1)
    dd_row_count = max(wb["Source_Chart_DD_Wide"].max_row - 1, 1)
    row_count = max(idx_row_count, dd_row_count)
    for row_index in range(6, 6 + row_count):
        source_data_row_formula = "ROW()-5"
        ws.cell(
            row=row_index,
            column=17,
            value=f'=IFERROR(INDEX(Source_Chart_IDX_Wide!$A:$A,{source_data_row_formula}+1),"")',
        )
        ws.cell(
            row=row_index,
            column=21,
            value=f'=IFERROR(INDEX(Source_Chart_DD_Wide!$A:$A,{source_data_row_formula}+1),"")',
        )
        idx_series_refs = ["$F$6", "$F$7"]
        dd_series_refs = idx_series_refs
        for offset, series_id_ref in enumerate(idx_series_refs, start=18):
            ws.cell(
                row=row_index,
                column=offset,
                value=_performance_chart_formula(
                    series_id_ref,
                    "Source_Chart_IDX_Wide",
                    "Source_Chart_IDX_Wide!$1:$1",
                    source_data_row_formula,
                ),
            )
        for offset, series_id_ref in enumerate(dd_series_refs, start=22):
            ws.cell(
                row=row_index,
                column=offset,
                value=_performance_chart_formula(
                    series_id_ref,
                    "Source_Chart_DD_Wide",
                    "Source_Chart_DD_Wide!$1:$1",
                    source_data_row_formula,
                ),
            )

    for name, cell_ref in {
        "Calc_Category_Selected_Portfolio": "B6",
        "Calc_Category_Selected_Variant": "B7",
        "Calc_Category_Selected_Period": "B8",
        "Calc_Category_Available_Count": "B13",
        "Calc_Category_Status": "B14",
        "Calc_Category_Chart_Selected": "$D$5:$F$7",
        "Calc_Category_Index_Data": f"$Q$5:$S${5 + row_count}",
        "Calc_Category_Drawdown_Data": f"$U$5:$W${5 + row_count}",
    }.items():
        _add_named_range(wb, name, ws.title, cell_ref)
    _add_formula_named_range(
        wb,
        "List_Category_Available_Series",
        f"=OFFSET('Calc_Category'!$K${list_first_row},0,0,MAX(1,COUNTIF('Calc_Category'!$K${list_first_row}:$K${list_last_row},\"<>\")),1)",
    )
    _add_formula_named_range(
        wb,
        "List_Category_Benchmark_Series",
        f"=OFFSET('Calc_Category'!$N${benchmark_first_row},0,0,MAX(1,COUNTIF('Calc_Category'!$N${benchmark_first_row}:$N${benchmark_last_row},\"<>\")),1)",
    )

    ws.freeze_panes = "A5"
    for column_letter, width in {
        "A": 24,
        "B": 38,
        "D": 16,
        "E": 28,
        "F": 22,
        "J": 14,
        "K": 28,
        "L": 22,
        "N": 28,
        "O": 22,
        "Q": 14,
        "R": 14,
        "S": 14,
        "U": 14,
        "V": 14,
        "W": 14,
    }.items():
        ws.column_dimensions[column_letter].width = width
    ws.column_dimensions["J"].hidden = True


def _build_category_sheet(wb: Workbook, defaults: dict[str, str], category_row_limit: int) -> None:
    ws = wb.create_sheet("Category")
    ws["A1"] = "Category"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Category analysis (REAL only) with its own editable controls."

    status_headers = ["Portfolio", "Variant", "Period", "Chart category", "Benchmark", "Data through"]
    status_formulas = [
        defaults["selected_category_portfolio"],
        "REAL",
        defaults["selected_category_period"],
        defaults.get("selected_category_chart", ""),
        defaults.get("selected_category_benchmark", ""),
        '=IFERROR(INDEX(Source_Build_Date_Max,1),"")',
    ]
    for column_index, value in enumerate(status_headers, start=1):
        ws.cell(row=4, column=column_index, value=value).font = Font(bold=True)
    for column_index, value in enumerate(status_formulas, start=1):
        ws.cell(row=5, column=column_index, value=value)
    ws["F5"].number_format = "yyyy-mm-dd"
    ws["A6"] = "Edit category selections here. This sheet is the only source for category controls."

    for range_name, cell_ref in CATEGORY_CONTROL_SOURCE_CELLS.items():
        _add_named_range(wb, range_name, ws.title, cell_ref)
    _add_list_validation(ws, "A5", "=List_Category_Portfolio")
    _add_list_validation(ws, "B5", '="REAL"')
    _add_list_validation(ws, "C5", "=List_Period")
    _add_list_validation(ws, "D5", "=List_Category_Available_Series")
    _add_list_validation(ws, "E5", "=List_Category_Benchmark_Series")

    ws["A7"] = "Category KPI comparison"
    ws["A7"].font = Font(bold=True)
    kpi_headers = ["Category", "Series_ID", "Period", "Return", "CAGR", "Vol", "Sharpe", "Max_DD", "Calmar"]
    for column_index, header in enumerate(kpi_headers, start=1):
        ws.cell(row=8, column=column_index, value=header).font = Font(bold=True)
    kpi_start_row = 9
    kpi_end_row = kpi_start_row + category_row_limit - 1
    for row_index in range(kpi_start_row, kpi_end_row + 1):
        calc_row = row_index - 3
        series_id_ref = f"$B{row_index}"
        ws[f"A{row_index}"] = _selected_label_formula(f"'Calc_Category'!K{calc_row}", series_id_ref)
        ws[f"B{row_index}"] = f"='Calc_Category'!L{calc_row}"
        ws[f"C{row_index}"] = f'=IF({series_id_ref}="","",Calc_Category_Selected_Period)'
        ws[f"D{row_index}"] = _category_kpi_lookup_formula(series_id_ref, "Source_KPI_Return_Total")
        ws[f"E{row_index}"] = _category_kpi_lookup_formula(series_id_ref, "Source_KPI_CAGR")
        ws[f"F{row_index}"] = _category_kpi_lookup_formula(series_id_ref, "Source_KPI_Vol")
        ws[f"G{row_index}"] = _category_kpi_lookup_formula(series_id_ref, "Source_KPI_Sharpe")
        ws[f"H{row_index}"] = _category_kpi_lookup_formula(series_id_ref, "Source_KPI_Max_DD")
        ws[f"I{row_index}"] = _category_kpi_lookup_formula(series_id_ref, "Source_KPI_Calmar")

    period_header_row = kpi_end_row + 2
    period_start_row = period_header_row + 2
    period_end_row = period_start_row + category_row_limit - 1
    ws[f"A{period_header_row}"] = "Category Period Returns"
    ws[f"A{period_header_row}"].font = Font(bold=True)
    period_headers = ["Category", "Series_ID", "30D", "YTD", "1Y", "Since_Start"]
    for column_index, header in enumerate(period_headers, start=1):
        ws.cell(row=period_header_row + 1, column=column_index, value=header).font = Font(bold=True)
    for row_index in range(period_start_row, period_end_row + 1):
        calc_row = row_index - period_start_row + 6
        series_id_ref = f"$B{row_index}"
        ws[f"A{row_index}"] = _selected_label_formula(f"'Calc_Category'!K{calc_row}", series_id_ref)
        ws[f"B{row_index}"] = f"='Calc_Category'!L{calc_row}"
        ws[f"C{row_index}"] = _category_period_lookup_formula(series_id_ref, "Source_Period_30D")
        ws[f"D{row_index}"] = _category_period_lookup_formula(series_id_ref, "Source_Period_YTD")
        ws[f"E{row_index}"] = _category_period_lookup_formula(series_id_ref, "Source_Period_1Y")
        ws[f"F{row_index}"] = _category_period_lookup_formula(series_id_ref, "Source_Period_Since_Start")

    chart_section_row = period_end_row + 2
    chart_header_row = chart_section_row + 1
    chart_data_start_row = chart_header_row + 1
    ws[f"A{chart_section_row}"] = "Index development"
    ws[f"A{chart_section_row}"].font = Font(bold=True)
    category_chart_headers = {
        f"A{chart_header_row}": "Date",
        f"B{chart_header_row}": _selected_label_formula("'Calc_Category'!E6", "'Calc_Category'!F6"),
        f"C{chart_header_row}": _selected_label_formula("'Calc_Category'!E7", "'Calc_Category'!F7"),
        f"G{chart_header_row}": "Date",
        f"H{chart_header_row}": _selected_label_formula("'Calc_Category'!E6", "'Calc_Category'!F6"),
        f"I{chart_header_row}": _selected_label_formula("'Calc_Category'!E7", "'Calc_Category'!F7"),
    }
    for cell_ref, value in category_chart_headers.items():
        ws[cell_ref] = value
        ws[cell_ref].font = Font(bold=True)

    idx_row_count = max(wb["Source_Chart_IDX_Wide"].max_row - 1, 1)
    dd_row_count = max(wb["Source_Chart_DD_Wide"].max_row - 1, 1)
    row_count = max(idx_row_count, dd_row_count)
    category_chart_max_row = chart_header_row + row_count
    for row_index in range(chart_data_start_row, category_chart_max_row + 1):
        calc_row = row_index - chart_data_start_row + 6
        ws[f"A{row_index}"] = f"='Calc_Category'!Q{calc_row}"
        ws[f"B{row_index}"] = f"='Calc_Category'!R{calc_row}"
        ws[f"C{row_index}"] = f"='Calc_Category'!S{calc_row}"
        ws[f"G{row_index}"] = f"='Calc_Category'!U{calc_row}"
        ws[f"H{row_index}"] = f"='Calc_Category'!V{calc_row}"
        ws[f"I{row_index}"] = f"='Calc_Category'!W{calc_row}"

    ws[f"G{chart_section_row}"] = "Drawdown"
    ws[f"G{chart_section_row}"].font = Font(bold=True)

    _build_performance_chart(ws, "Category index development", f"A{chart_section_row + 8}", 1, 2, 3, chart_header_row, category_chart_max_row)
    _build_performance_chart(ws, "Category drawdown", f"J{chart_section_row + 8}", 7, 8, 9, chart_header_row, category_chart_max_row)

    for row_index in range(kpi_start_row, kpi_end_row + 1):
        for column_letter in ["D", "E", "F", "H"]:
            ws[f"{column_letter}{row_index}"].number_format = "0.00%"
        for column_letter in ["G", "I"]:
            ws[f"{column_letter}{row_index}"].number_format = "0.00"
    for row_index in range(period_start_row, period_end_row + 1):
        for column_letter in ["C", "D", "E", "F"]:
            ws[f"{column_letter}{row_index}"].number_format = "0.00%"
    for row_index in range(chart_data_start_row, category_chart_max_row + 1):
        for column_letter in ["B", "C", "H", "I"]:
            ws[f"{column_letter}{row_index}"].number_format = "0.00%"
        ws[f"A{row_index}"].number_format = "yyyy-mm-dd"
        ws[f"G{row_index}"].number_format = "yyyy-mm-dd"

    for column_letter, width in {
        "A": 14,
        "B": 26,
        "C": 22,
        "D": 14,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 12,
        "I": 12,
        "J": 12,
    }.items():
        ws.column_dimensions[column_letter].width = width
    ws.column_dimensions["B"].hidden = True
    ws.freeze_panes = "A4"


def build_dashboard_workbook(source_path: Path, output_path: Path) -> None:
    sheets, found_sheets = _load_dashboard_data(source_path)
    logging.info("Dashboard workbook input path: %s", source_path)
    logging.info("Dashboard workbook output path: %s", output_path)
    logging.info("Dashboard workbook source sheets found: %s", ", ".join(found_sheets))

    series_catalog = _build_series_catalog(sheets["KPI_Summary"])
    dashboard_config_map = _config_map(sheets["Dashboard_Config"])

    main_portfolios = sorted(
        {
            portfolio
            for portfolio in series_catalog.loc[series_catalog["Is_Main_Portfolio"], "Portfolio_Name"].tolist()
            if portfolio
        }
    )
    compare_series = sorted(
        {
            display_name
            for display_name in series_catalog.loc[series_catalog["Is_Compare_Eligible"], "Display_Name"].tolist()
            if display_name
        }
    )
    main_series_rows = (
        series_catalog.loc[series_catalog["Is_Main_Portfolio"], ["Portfolio_Name", "Variant", "Display_Name", "Series_ID"]]
        .sort_values(["Portfolio_Name", "Variant", "Display_Name", "Series_ID"])
        .values.tolist()
    )
    compare_lookup_rows = (
        series_catalog.loc[
            series_catalog["Is_Compare_Eligible"],
            ["Display_Name", "Series_ID", "Series_Type", "Portfolio_Name", "Variant"],
        ]
        .sort_values(["Display_Name", "Series_ID"])
        .values.tolist()
    )
    category_catalog = series_catalog.loc[
        series_catalog["Is_Category"] & (series_catalog["Variant"] == "REAL")
    ].copy()
    category_portfolios = sorted(
        {
            portfolio
            for portfolio in category_catalog["Portfolio_Name"].tolist()
            if portfolio
        }
    )
    category_rows = (
        category_catalog.loc[:, ["Portfolio_Name", "Display_Name", "Series_ID", "Variant"]]
        .sort_values(["Portfolio_Name", "Display_Name", "Series_ID", "Variant"])
        .values.tolist()
    )
    benchmark_rows = (
        series_catalog.loc[series_catalog["Series_Type"] == "BM", ["Display_Name", "Series_ID"]]
        .sort_values(["Display_Name", "Series_ID"])
        .values.tolist()
    )

    default_portfolio = _choose_default_portfolio(main_portfolios, dashboard_config_map)
    default_period = _choose_default_period(dashboard_config_map)
    category_default_portfolio = (
        default_portfolio
        if default_portfolio in category_portfolios
        else (category_portfolios[0] if category_portfolios else "")
    )
    defaults = {
        "selected_portfolio": default_portfolio,
        "selected_variant": "REAL",
        "selected_period": default_period,
        "selected_category_portfolio": category_default_portfolio,
        "selected_category_period": default_period,
        "selected_category_chart": _choose_default_category_selection(category_rows, category_default_portfolio, 1),
        "selected_category_benchmark": _choose_default_benchmark(benchmark_rows),
    }
    category_helper_row_limit = max(len(category_rows), CATEGORY_FILTER_MAX_ROWS, 1)
    category_visible_row_limit = max(min(len(category_rows), CATEGORY_VISIBLE_MAX_ROWS), 1)
    benchmark_row_limit = max(len(benchmark_rows), 1)

    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    _build_overview_source_sheets(wb, sheets)
    _build_performance_source_sheets(wb, sheets)
    _build_structure_source_sheets(wb, sheets)
    _build_overview_sheet(wb, defaults)
    _build_calc_main_sheet(wb)
    _build_performance_sheet(wb)
    _build_structure_sheet(wb)
    _build_lists_sheet(
        wb,
        main_portfolios,
        compare_series,
        category_portfolios,
        main_series_rows,
        compare_lookup_rows,
        category_rows,
        benchmark_rows,
    )
    _build_calc_category_sheet(wb, category_helper_row_limit, benchmark_row_limit)
    _build_category_sheet(wb, defaults, category_visible_row_limit)
    _build_control_sheet(wb, defaults, source_path, output_path, found_sheets)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logging.info("Dashboard workbook written: %s", output_path)


def run(
    dashboard_data_path: str | Path | None = None,
    dashboard_workbook_output_path: str | Path | None = None,
) -> None:
    _configure_logging()
    source_path = Path(dashboard_data_path or config.DASHBOARD_DATA_OUTPUT_PATH)
    output_path = Path(dashboard_workbook_output_path or config.DASHBOARD_WORKBOOK_OUTPUT_PATH)
    build_dashboard_workbook(source_path, output_path)


if __name__ == "__main__":
    args = _parse_args()
    run(
        dashboard_data_path=args.input_path,
        dashboard_workbook_output_path=args.output_path,
    )
