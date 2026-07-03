"""Bygg om pipelinens ``transaktioner.xlsx`` från en rå Nordnet-export.

Detta skript är det ENDA reproducerbara spåret av Pass 2-datafixen. Själva
``transaktioner.xlsx`` versionshanteras inte (den ligger i den maskinlokala
dataroten och är .gitignore:ad via ``data/*.xlsx``-mönstret på andra maskiner /
utanför repot), så logiken som producerar den bor här i stället.

Bakgrund
--------
Den tidigare transaktionsfilen saknade BYTESKÖP/BYTESSÄLJ-rader, vilket gav
negativa ackumulerade innehav (sålt mer än köpt) och därmed felaktig REAL-serie.
Den kompletta exporten ``transactions_full_export.xlsx`` innehåller alla typer.

Transaktionstyp → tecken på Antal
---------------------------------
Pipelinen (``src.portfolio._txn_sign``) läser tecknet från Transaktionstyp:
``K*`` → +1, ``S*`` → −1. Antal skrivs alltid som positiv magnitud. Därför
relabelas de positionspåverkande typerna till KÖPT/SÅLT enligt:

    POSITIV  (→ "KÖPT"):  KÖPT, BYTESKÖP, EM INLÄGG VP
    NEGATIV  (→ "SÅLT"):  SÅLT, BYTESSÄLJ, RENSNING UTTAG VP

Alla övriga typer utesluts HELT ur filen (de är inga positionsändringar och får
inte förorena vare sig positioner eller kassaflöde):

    UTDELNING, KÄLLSKATT, KAP RÄNTA, INSÄTTNING, KÖP VALUTA, SÄLJ VALUTA, UTTAG

Rader utan ISIN (t.ex. EM INLÄGG/RENSNING på "ANSÖKAN ...") hoppas över – de
nettar ändå till 0. ``KÄLLSKATT`` börjar på "K" och har både ISIN och Antal;
utan uteslutning skulle den felaktigt räknas som ett köp – därför är
typuteslutningen kritisk.

Kassaflödes-neutraliserad seeding av PA:s ingående balans
---------------------------------------------------------
PA hade innehav vid årsskiftet 2023/24 som köptes före exportfönstret. De seedas
som öppningsrader. En naiv seed FÖRE fönstret med Belopp=0 fungerar INTE, eftersom
flera fonders Yahoo-priser börjar först 2024-01-08/09 – deras värde "dyker upp"
mitt i fönstret utan motverkande kassaflöde och skapar en falsk dagsavkastning
(observerat +92 %). I stället seedas varje fond på sitt FÖRSTA prisdatum med

    Belopp = -(antal × pipelinens egen kurs_base den dagen), Valuta = SEK

så att marknadsvärdet som tillkommer exakt motsvaras av ett kassaflöde och TWR
nollar ingången. Metoden matchar masterns beprövade IB-rader (korsvaliderad:
0,00 % avvikelse för 4 av 5 överlappande fonder; USD-fonden LU0211331839 skiljer
~1,8 % pga FX/NAV-timing men neutraliseras exakt eftersom Belopp beräknas från
motorns egen kurs). Antalen nedan är oförändrade från Nordnets årsbesked.

EGEN får INGEN seed – dess export är självbärande inom fönstret.

Körning
-------
    python tools/rebuild_transaktioner.py [utdata.xlsx]

Bygger till ``utdata.xlsx`` (default ``data/transaktioner_rebuilt.xlsx``, som är
regenererbar). Skriptet skriver ALDRIG över den skarpa filen automatiskt – ta
backup och kopiera manuellt efter granskning:

    cp <config PATH_TRANSAKTIONER> <...>.bak     # om ingen backup finns
    cp data/transaktioner_rebuilt.xlsx <config PATH_TRANSAKTIONER>
"""

from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Gör "src"-paketet importerbart oavsett arbetskatalog.
REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

import pandas as pd  # noqa: E402

from src import config  # noqa: E402
from src.bootstrap import init_ssl  # noqa: E402
from src.io_excel import load_inputs  # noqa: E402
from src.portfolio import _fx_tickers_for_assets, _prices_to_base  # noqa: E402
from src.prices import download_adj_close  # noqa: E402

# Rå Nordnet-export (Affärsdag 2024-01-31 → …, båda depåerna).
RAW = Path(config.PATH_TRANSAKTIONER).with_name("transactions_full_export.xlsx")
RAW_SHEET = "transactions-and-notes-export ("

POS = {"KÖPT", "BYTESKÖP", "EM INLÄGG VP"}        # → "KÖPT"  (positiv positionseffekt)
NEG = {"SÅLT", "BYTESSÄLJ", "RENSNING UTTAG VP"}  # → "SÅLT"  (negativ positionseffekt)

PA_DEPA = 10110120  # Depå-ID för portfölj PA

# PA:s ingående balans per årsskiftet 2023/24 (Nordnets årsbesked). Antal fastställt,
# datum + Belopp beräknas av skriptet från första prisdatum × pipelinens kurs.
PA_SEED_QTY = {
    "SE0016830749": 778.3073,
    "SE0004636447": 124.2824,
    "SE0004452118": 3400.9679,
    "LU0211331839": 2839.816,
    "LU0637346080": 648.637,
    "SE0007871488": 390.9547,
    "LU2678175618": 1340.7725,
    "SE0000429789": 121.693476,
}

# Ny Mapping-rad för China Stars (byte-par, ticker bekräftad av ägaren).
# Kolumnordning: ISIN, Name, Yahoo_Ticker, Price_Source, Instrument_Type,
#                Price_Currency, Category, Geography
NEW_MAPPING_ROWS = [
    ["LU1608617111", "AGCM China Stars RC1 SEK", "0P0001BIYV.ST", "Yahoo Finance",
     "Fond", "SEK", "Tillväxtmarknader", "Asien"],
]

MAPPING_HEADER = ["ISIN", "Name", "Yahoo_Ticker", "Price_Source",
                  "Instrument_Type", "Price_Currency", "Category", "Geography"]


def load_sheet(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows[0], rows[1:]


def compute_seed_rows(ci, name_by_isin):
    """Beräkna förskjutna, neutraliserade seed-rader för PA."""
    init_ssl()
    tables = load_inputs(config.PATH_TRANSAKTIONER, config.PATH_FONDER)
    mapping = tables["mapping"].copy()
    mapping["ISIN"] = mapping["ISIN"].astype(str).str.strip()
    isin_to_ticker = mapping.set_index("ISIN")["Yahoo_Ticker"].to_dict()

    seed_isins = list(PA_SEED_QTY)
    tickers = [isin_to_ticker[i] for i in seed_isins]
    fx = _fx_tickers_for_assets(tickers, mapping, config.BASE_CURRENCY)
    prices = download_adj_close(
        tickers=sorted(set(tickers) | set(fx)),
        start_date=pd.to_datetime(tables["portfolio_metadata"]["Index_Start_Date"]).min(),
        end_date=None,
        forward_fill=config.FORWARD_FILL,
        fx_tickers=fx,
    )
    price_base = _prices_to_base(prices, tickers, mapping, config.BASE_CURRENCY)

    rows = []
    for isin in seed_isins:
        qty = PA_SEED_QTY[isin]
        ticker = isin_to_ticker[isin]
        series = price_base[ticker]
        first_date = series.first_valid_index()
        if first_date is None:
            raise SystemExit(f"Saknar prisdata för seed-fond {isin} ({ticker})")
        belopp = -(qty * float(series.loc[first_date]))
        seed_date = first_date.to_pydatetime()

        row = [None] * len(ci)
        row[ci["Bokföringsdag"]] = seed_date
        row[ci["Affärsdag"]] = seed_date
        row[ci["Likviddag"]] = seed_date
        row[ci["Depå"]] = PA_DEPA
        row[ci["Transaktionstyp"]] = "KÖPT"
        row[ci["Värdepapper"]] = name_by_isin.get(isin, isin)
        row[ci["ISIN"]] = isin
        row[ci["Antal"]] = float(qty)
        row[ci["Valuta"]] = "SEK"
        row[ci["Belopp"]] = float(belopp)  # -(antal×kurs) → TWR nollar ingången
        row[ci["Transaktionstext"]] = "IB (seed opening balance, neutraliserad)"
        rows.append(row)
        print(f"  seed {isin} {ticker:14s} {seed_date.date()} Belopp={belopp:.2f}")
    return rows


def transform_position_rows(ci, raw_rows):
    """Behåll KÖPT/SÅLT/BYTES*, relabela byten, uteslut övrigt."""
    out, relabel = [], {"BYTESKÖP": 0, "BYTESSÄLJ": 0}
    excl_type = excl_noisin = 0
    for r in raw_rows:
        typ, isin = r[ci["Transaktionstyp"]], r[ci["ISIN"]]
        if typ not in POS and typ not in NEG:
            excl_type += 1
            continue
        if isin is None or str(isin).strip() == "":
            excl_noisin += 1
            continue
        row = list(r)
        row[ci["Transaktionstyp"]] = "KÖPT" if typ in POS else "SÅLT"
        if typ in ("BYTESKÖP", "BYTESSÄLJ"):
            relabel[typ] += 1
            if not str(row[ci["Transaktionstext"]] or "").strip():
                row[ci["Transaktionstext"]] = typ  # bevara ursprungstyp för spårbarhet
        out.append(row)
    print(f"  relabel={relabel} uteslutna_typ={excl_type} uteslutna_utan_ISIN={excl_noisin}")
    return out


def add_table(wb, sheet_name, header, data_rows, table_name):
    ws = wb.create_sheet(sheet_name)
    ws.append(list(header))
    for row in data_rows:
        ws.append(list(row))
    ref = f"A1:{get_column_letter(len(header))}{len(data_rows) + 1}"
    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True,
        showFirstColumn=False, showLastColumn=False, showColumnStripes=False,
    )
    ws.add_table(tab)
    return ref


def main():
    out_path = Path(sys.argv[1]) if len(sys.argv) > 1 else REPO_ROOT / "data" / "transaktioner_rebuilt.xlsx"

    raw_hdr, raw_rows = load_sheet(RAW, RAW_SHEET)
    cols = list(raw_hdr)  # 30 kolumner, exakt Nordnet-ordning
    ci = {name: i for i, name in enumerate(cols)}

    name_by_isin = {}
    for r in raw_rows:
        isin = r[ci["ISIN"]]
        if isin is not None:
            name_by_isin.setdefault(str(isin).strip(), r[ci["Värdepapper"]])

    print("Beräknar neutraliserade seed-rader …")
    seed_rows = compute_seed_rows(ci, name_by_isin)
    print("Transformerar positionsrader …")
    pos_rows = transform_position_rows(ci, raw_rows)
    out_rows = seed_rows + pos_rows
    print(f"Datarader ut: {len(out_rows)} ({len(seed_rows)} seed + {len(pos_rows)} positioner)")

    # Bevarade tabeller läses som VÄRDEN (exakt vad pipelinen redan konsumerar),
    # vilket materialiserar Mappings Name-formler och undviker openpyxl:s cache-fälla.
    preserved = {s: load_sheet(config.PATH_TRANSAKTIONER, s)
                 for s in ("Benchmarks", "Mapping", "Portfolio_Metadata")}

    assert list(preserved["Mapping"][0]) == MAPPING_HEADER, preserved["Mapping"][0]
    existing = {str(r[0]).strip() for r in preserved["Mapping"][1]}
    for nr in NEW_MAPPING_ROWS:
        if nr[0] not in existing:
            preserved["Mapping"][1].append(nr)
            print(f"  Mapping-rad tillagd: {nr[0]} {nr[2]}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    add_table(wb, "Transactions", cols, out_rows, "Transactions")
    for sheet in ("Benchmarks", "Mapping", "Portfolio_Metadata"):
        add_table(wb, sheet, preserved[sheet][0], preserved[sheet][1], sheet)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Skrev: {out_path}")
    print("OBS: kopiera manuellt till config PATH_TRANSAKTIONER efter granskning (ta backup först).")


if __name__ == "__main__":
    main()
